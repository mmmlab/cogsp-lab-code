# unit: pixels on the screen
from psychopy import visual, core, event, monitors,data
from psychopy import gui
from pyglet.window import key
#import xlwt
import openpyxl as pyxl
import os
import random
import math
s = False
global win
result=[]


# def write_file(file,sq_contrast,apertWidth):
#     global result
#     book = xlwt.Workbook(encoding="utf-8")
#     sheet1 = book.add_sheet("Sheet 1")
#     sheet1.write(0, 0, "ExperimentName:%s"%name)
#     sheet1.write(1, 0, "Student Initial:%s"%ID)
#     sheet1.write(2, 0, "SessionNumber:%d"%session)
#     sheet1.write(3, 0, "Trail#")
#     sheet1.write(3, 1, "barContrast")
#     sheet1.write(3, 2, "ApertureWidth")
#     sheet1.write(3, 3, "response")
#     for i in range(0,len(result)):
#         sheet1.write(i+4, 0, i+1)
#         sheet1.write(i+4, 1, sq_contrast)
#         sheet1.write(i+4, 2, apertWidth)
#         sheet1.write(i+4, 3, result[i])
#     file+='.xls'
#     book.save(file) 

# this function writes output to a file
def write_file(filename,square_contrast,aperture_width):
    global result
    book = pyxl.Workbook()
    ws = book.active
    ws.title = "Sheet 1"
    # write experiment information
    ws.cell(row=1,column=1).value = "ExperimentName:%s"%name
    ws.cell(row=2,column=1).value = "SubjectID:%s"%ID
    ws.cell(row=3,column=1).value = "SessionNumber:%s"%session
    # write column headers
    headers = ['Trial','BarContrast','ApertureWidth','Response']
    for col,header in enumerate(headers):
        ws.cell(row=4,column=col+1).value = header
    # write per-trial data
    for r in range(0,len(result)):
        ws.cell(r+5, 1).value =  r+1
        ws.cell(r+5, 2).value = square_contrast
        ws.cell(r+5, 3).value = aperture_width
        ws.cell(r+5, 4).value = result[r]
    # create directory and/or save file
    if not os.path.exists('data'):
        os.makedirs('data')
    book.save('data/'+filename+'.xlsx')


def show_rec(contrast,apertWidth):
    global win
    global result
    square_len=210
    square_width=10
    square_color=[i*contrast for i in [255.0,255.0,255.0]] # contrast of the bars
    rm=10; # radius of motion
    r=square_len/2 # radium for the occluder
    # set square motion
    angular_speed=0.11 # angular speed, radian per second
    duration=3 # seconds
    # set square occluder shape,size,contrast
    occ_size=100
    occ_contrast=0.5
    # set circular background shape,size,contrast
    cir_size=80
    cir_contrast=occ_contrast
    # set aperture shape,loc, size,contrast
    aperture_len=2*(r-occ_size/2)
    aperture_height=apertWidth*aperture_len
    aperture_contrast=0.1
    #create stimuli: 1 square,4 apertures and 4 square occluders and 4 circular background
    # 4 occluders
    square_loc_initial=[0,0]
    occ_locsx=[r,r,-r,-r]
    occ_locsy=[r,-r,-r,r]
    occ1=visual.Rect(win,units='pix',fillColor=1,width=occ_size,height=occ_size,contrast=occ_contrast,pos=(occ_locsx[0],occ_locsy[0]));
    occ2=visual.Rect(win,units='pix',fillColor=1,width=occ_size,height=occ_size,contrast=occ_contrast,pos=(occ_locsx[1],occ_locsy[1]));
    occ3=visual.Rect(win,units='pix',fillColor=1,width=occ_size,height=occ_size,contrast=occ_contrast,pos=(occ_locsx[2],occ_locsy[2]));
    occ4=visual.Rect(win,units='pix',fillColor=1,width=occ_size,height=occ_size,contrast=occ_contrast,pos=(occ_locsx[3],occ_locsy[3]));
    # 4 circular background
    cir_locsx=occ_locsx
    cir_locsy=occ_locsy
    cir1=visual.Circle(win,units='pix',fillColor=1,radius=cir_size,contrast=cir_contrast,pos=(cir_locsx[0],cir_locsy[0]));
    cir2=visual.Circle(win,units='pix',fillColor=1,radius=cir_size,contrast=cir_contrast,pos=(cir_locsx[1],cir_locsy[1]));
    cir3=visual.Circle(win,units='pix',fillColor=1,radius=cir_size,contrast=cir_contrast,pos=(cir_locsx[2],cir_locsy[2]));
    cir4=visual.Circle(win,units='pix',fillColor=1,radius=cir_size,contrast=cir_contrast,pos=(cir_locsx[3],cir_locsy[3]));
    # 4 apertures
    aperture_locsx=[0,r,0,-r]
    aperture_locsy=[r,0,-r,0]
    apert1=visual.Rect(win,units='pix',lineColorSpace='rgb255',lineColor=[x*aperture_contrast for x in [255,255,255]],fillColorSpace='rgb255',fillColor=[x*aperture_contrast for x in [255,255,255]],contrast=aperture_contrast,width=aperture_len,height=aperture_height,pos=(aperture_locsx[0],aperture_locsy[0]))
    apert2=visual.Rect(win,units='pix',lineColorSpace='rgb255',lineColor=[x*aperture_contrast for x in [255,255,255]],fillColorSpace='rgb255',fillColor=[x*aperture_contrast for x in [255,255,255]],contrast=aperture_contrast,width=aperture_height,height=aperture_len,pos=(aperture_locsx[1],aperture_locsy[1]))
    apert3=visual.Rect(win,units='pix',lineColorSpace='rgb255',lineColor=[x*aperture_contrast for x in [255,255,255]],fillColorSpace='rgb255',fillColor=[x*aperture_contrast for x in [255,255,255]],contrast=aperture_contrast,width=aperture_len,height=aperture_height,pos=(aperture_locsx[2],aperture_locsy[2]))
    apert4=visual.Rect(win,units='pix',lineColorSpace='rgb255',lineColor=[x*aperture_contrast for x in [255,255,255]],fillColorSpace='rgb255',fillColor=[x*aperture_contrast for x in [255,255,255]],contrast=aperture_contrast,width=aperture_height,height=aperture_len,pos=(aperture_locsx[3],aperture_locsy[3]))
    # draw
    cir1.draw();cir2.draw(); cir3.draw(); cir4.draw();apert1.draw();apert2.draw();apert3.draw();apert4.draw();occ1.draw();occ2.draw(); occ3.draw(); occ4.draw();
    
    # show motion
    for i in range(0,200): # 200 frames
        # square_loc=(rm*math.cos(angular_speed*i)+square_loc_initial[0],rm*math.sin(angular_speed*i)+square_loc_initial[1]);
        square_loc=(rm*math.cos(angular_speed*i)+square_loc_initial[0],rm*math.sin(angular_speed*i)+square_loc_initial[1]);
        square=visual.Rect(win,units='pix',lineColorSpace='rgb255',lineColor=[255,255,255],opacity=1,lineWidth=square_width,width=square_len,height=square_len,contrast=square_contrast,pos=(square_loc));   
       # square.colorSpace='rgb'
        cir1.draw();cir2.draw(); cir3.draw(); cir4.draw();apert1.draw();apert2.draw();apert3.draw();apert4.draw();square.draw();occ1.draw();occ2.draw(); occ3.draw(); occ4.draw();
        if i==199:
            showbox2.draw()
        win.update();
        core.wait(duration/200)
    while True:
        keylist=event.waitKeys()
        if 'escape' in keylist:
            print('exit by the users')
            result.append('NULL')
            mywin.close()
            core.quit()
            break 
        elif '1' in keylist:
            result.append(int(1))
            break
        elif '2' in keylist:
            result.append(int(2))
            break
        elif '3' in keylist:
            result.append(int(3))
            break
        else:
            # print 'incorrect key. ..Press 1 for incoherent, 2 for intermediate and 3 for coherent'
            cir1.draw();cir2.draw(); cir3.draw(); cir4.draw();apert1.draw();apert2.draw();apert3.draw();apert4.draw();square.draw();occ1.draw();occ2.draw(); occ3.draw(); occ4.draw();
            showbox2.setText('Incorrect key. Press 1 for incoherent, 2 for intermediate and 3 for coherent')
            showbox2.draw()
            win.update()
            continue
    

def show_instruction():
     global win
     instr=visual.SimpleImageStim(win, image='instruction.png', units='', pos=(0.0, 0.0), flipHoriz=False, flipVert=False, name=None, autoLog=None)
     instrStim = visual.BufferImageStim(win, stim = [instr])
     instrStim.draw()
     win.flip()
     while True:
        if len(event.getKeys())>0: break
     event.clearEvents()


#------------------------------------------------------------------------------#
################ This is the start of the actual script ########################
#------------------------------------------------------------------------------#          
# show GUI
myDlg = gui.Dlg(title="Motion experiment:shape of aperture",size=(1, 1))
myDlg.addField('ExperimentName:','Shape of Aperture')
myDlg.addField('Student Initial:','xh')
myDlg.addField('SessionNumber:', '001')
myDlg.addField('barContrast:', 1.0)
myDlg.addField('aperture width relative to its length(from 0.2 to 0.7):', 0.5)
myDlg.addField('NumberofTrials',5)
myDlg.show()
if myDlg.OK:
    thisInfo = myDlg.data
    name='Motion3'
    ID=thisInfo[1]
    session=thisInfo[2]
    barContrast=thisInfo[3]
    apertWidth=thisInfo[4]
    if apertWidth>0.7:
        apertWidth=0.7
    if apertWidth<0.2:
        apertWIDTH=0.2
    trials=thisInfo[5]
    file=''
    file+=name
    # file+='_Bar'+str(barContrast)
    # file+='_AperWidth'+str(apertWidth)
    file+='_'+ID
    file+='_Session'+str(session)
    file+='_'+data.getDateStr()
    print(file)
    keyState=key.KeyStateHandler()
    win = visual.Window( monitor="testMonitor", units="pix",allowGUI=True,fullscr=False)
    win.winHandle.push_handlers(keyState)
    show_instruction()
    # instruction between trials
    showbox1=visual.TextStim(win,
                        font='Arial',
                        height=0.09,
                        # font_color=[0,0,0],
                        # dpi=72,
                        # size=(1.8,0.6),
                        pos=(0.0,0), 
                        units='norm',
                        alignHoriz='center',
                        alignVert='center',
                        colorSpace='rgb255'
                        )
    global showbox2
    showbox2=visual.TextStim(win,
                        text='Response: 1 (incoherent), 2 (intermediate coherence), 3 (coherent)', 
                        font='Arial',
                        height=0.07,
                        # font_color=[0,0,0],
                        # dpi=72,
                        # size=(1.8,0.6),
                        pos=(0.0,-0.7), 
                        units='norm',
                        alignHoriz='center',
                        alignVert='center',
                        colorSpace='rgb255'
                        )
    square_contrast=barContrast;  # temporary, can be changed for future use
    for x in range(0,trials):
        if s==True:
            break;
                # show instruction between trials
        showbox1.setText('Trial% .0f: Hit any key to see the display' %(x+1))
        showbox2.setText('Response: 1 (incoherent), 2 (intermediate coherence), 3 (coherent)')
        showbox1.draw()
        win.flip()
        # wait for user input
        while True:
            if len(event.getKeys())>0: break
        event.clearEvents()
        show_rec(barContrast,apertWidth);
    write_file(file,barContrast,apertWidth)
else:
    print('user cancelled')

