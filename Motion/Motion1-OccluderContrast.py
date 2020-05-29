# units: pixels
from psychopy import visual, core, event, monitors,data
from psychopy import gui
from pyglet.window import key
#import xlwt
import openpyxl as pyxl
import os
import random
import math
import numpy


s = False
global win
result=[]

# this function writes output to a file
def write_file(filename,occ_contrast,square_contrast):
    global result
    book = pyxl.Workbook()
    ws = book.active
    ws.title = "Sheet 1"
    # write experiment information
    ws.cell(row=1,column=1).value = "ExperimentName:%s"%name
    ws.cell(row=2,column=1).value = "SubjectID:%s"%ID
    ws.cell(row=3,column=1).value = "SessionNumber:%s"%session
    # write column headers
    headers = ['Trial','SquareContrast','OccluderContrast','ContrastRatio','Response']
    for col,header in enumerate(headers):
        ws.cell(row=4,column=col+1).value = header
    # write per-trial data
    for r in range(0,len(result)):
        ws.cell(r+5, 1).value =  r+1
        ws.cell(r+5, 2).value = square_contrast
        ws.cell(r+5, 3).value = occ_contrast
        ws.cell(r+5, 4).value = square_contrast/occ_contrast
        ws.cell(r+5, 5).value = result[r]
    # create directory and/or save file
    if not os.path.exists('data'):
        os.makedirs('data')
    book.save('data/'+filename+'.xlsx')


def show_rec(square_contrast,occ_contrast):
    global win,result,showbox2
    # when occluder is rectangle
    square_len=210
    square_width=10
    square_color=[i*square_contrast for i in [255.0,255.0,255.0]] # contrast of the bars
    rm=10 # radius of motion
    r=square_len/2 # radium for the occluder
    # set square motion
    angular_speed=0.11 # angular speed, radian per second
    duration=3 # seconds
    # set occluder shape,size,contrast
    occ_size=100
    occ_contrast
    #create a window
    #create stimuli: square and occluders
    square_loc_initial=[0,0];
    occ_locsx=[r,r,-r,-r];
    occ_locsy=[r,-r,-r,r];
    occ1=visual.Rect(win,units='pix',fillColor=1,width=occ_size,height=occ_size,contrast=occ_contrast,pos=(occ_locsx[0],occ_locsy[0]));
    occ2=visual.Rect(win,units='pix',fillColor=1,width=occ_size,height=occ_size,contrast=occ_contrast,pos=(occ_locsx[1],occ_locsy[1]));
    occ3=visual.Rect(win,units='pix',fillColor=1,width=occ_size,height=occ_size,contrast=occ_contrast,pos=(occ_locsx[2],occ_locsy[2]));
    occ4=visual.Rect(win,units='pix',fillColor=1,width=occ_size,height=occ_size,contrast=occ_contrast,pos=(occ_locsx[3],occ_locsy[3]));
    occ1.draw();occ2.draw(); occ3.draw(); occ4.draw();
    # show motion
    for i in range(0,200): # 200 frames
        square_loc=(rm*math.cos(angular_speed*i)+square_loc_initial[0],rm*math.sin(angular_speed*i)+square_loc_initial[1]);
        square=visual.Rect(win,units='pix',lineColorSpace='rgb255',lineColor=[255,255,255],opacity=1,lineWidth=square_width,width=square_len,height=square_len,contrast=square_contrast,pos=(square_loc));   
        square.draw();occ1.draw();occ2.draw(); occ3.draw(); occ4.draw();
        if i==199:
            showbox2.draw()
        win.update()
        core.wait(duration/200);
    while True:
        keylist=event.waitKeys()
        if 'escape' in keylist:
            print('exit by the users')
            result.append('NULL')
            mywin.close()
            core.quit()
            break 
        elif '1' in keylist:
            result.append(int(1))
            break
        elif '2' in keylist:
            result.append(int(2))
            break
        elif '3' in keylist:
            result.append(int(3))
            break
        else:
            # print 'incorrect key. ..Press 1 for incoherent, 2 for intermediate and 3 for coherent'
            square.draw();occ1.draw();occ2.draw(); occ3.draw(); occ4.draw();
            showbox2.setText('Incorrect key.  Press 1 for incoherent, 2 for intermediate and 3 for coherent')
            showbox2.draw()
            win.update()
            continue


def show_instruction():
     global win
     instr=visual.SimpleImageStim(win, image='instruction.png', units='', pos=(0.0, 0.0), flipHoriz=False, flipVert=False, name=None, autoLog=None)
     instrStim = visual.BufferImageStim(win, stim = [instr])
     instrStim.draw()
     win.flip()
     while True:
        if len(event.getKeys())>0: break
     event.clearEvents()


#------------------------------------------------------------------------------#
################ This is the start of the actual script ########################
#------------------------------------------------------------------------------#     
# show GUI
myDlg = gui.Dlg(title="Motion interpretation Experiment",size=(1, 1))
myDlg.addField('ExperimentName:','Effect of occluder contrast')
myDlg.addField('Student Initial:','xh')
myDlg.addField('SessionNumber:', '001')
# myDlg.addField('Square Contrast:',0,choices=list(numpy.linspace(0,1,num=21)))
myDlg.addField('Bar Contrast(from 0 to 1):',1)
myDlg.addField('Occluders Contrast(from 0 to 1):',1)
myDlg.addField('NumberofTrials',5)
# myDlg.addFixedField('Contrast Ratio(square/occluder)',myDlg.data[3]/myDlg.data[4])
myDlg.show()

if myDlg.OK:
    thisInfo = myDlg.data
    name='Motion1'
    ID=thisInfo[1]
    session=thisInfo[2]
    square_contrast=thisInfo[3]
    occ_contrast=thisInfo[4]
    trials=thisInfo[5]
    file=''
    file+=name
    # file+='_Bar'+str(square_contrast)
    # file+='_Occ'+str(occ_contrast)
    file+='_'+ID
    file+='_Session'+str(session)
    file+='_'+data.getDateStr()
    print(file)
    keyState=key.KeyStateHandler()
    win = visual.Window( monitor="testMonitor", units="deg",allowGUI=True,fullscr=False)
    win.winHandle.push_handlers(keyState)
    show_instruction()
    # instruction between trials
    showbox1=visual.TextStim(win,
                        font='Arial',
                        height=0.09,
                        # font_color=[0,0,0],
                        # dpi=72,
                        # size=(1.8,0.6),
                        pos=(0.0,0), 
                        units='norm',
                        alignHoriz='center',
                        alignVert='center',
                        colorSpace='rgb255'
                        )
    global showbox2
    showbox2=visual.TextStim(win,
                        text='Response: 1 (incoherent), 2 (intermediate coherence), 3 (coherent)', 
                        font='Arial',
                        height=0.07,
                        # font_color=[0,0,0],
                        # dpi=72,
                        # size=(1.8,0.6),
                        pos=(0.0,-0.7), 
                        units='norm',
                        alignHoriz='center',
                        alignVert='center',
                        colorSpace='rgb255'
                        )
    for x in range(0,trials):
        if s==True:
            break;
        # show instruction between trials
        showbox1.setText('Trial % .0f: Hit any key to see the display' %(x+1))
        showbox2.setText('Response: 1 (incoherent), 2 (intermediate coherence), 3 (coherent)')
        showbox1.draw()
        win.flip()
        # wait for user input
        while True:
            if len(event.getKeys())>0: break
        event.clearEvents()
        # show display
        show_rec(square_contrast,occ_contrast);
        
    write_file(file,occ_contrast,square_contrast)
else:
    print('user cancelled')

