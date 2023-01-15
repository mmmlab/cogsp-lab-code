#visual_estm
#J Rubinstein   Oct 8, 2015
#based on.  
#E Kowler.  Aug 19, 2015

from psychopy import visual, core, gui, data, event, sound
from psychopy.tools.filetools import fromFile, toFile
import time, numpy, random, string, scipy
#import xlwt
import openpyxl as pyxl
import pyglet
import os

global mywin
global grating
global fixation
global blankletter
global centraltext
global xecc,yecc
global savetilt



# set up the menu for choice of conditions
myDlg = gui.Dlg(title="Visual Short Term Memory",size=(1, 1))
myDlg.addField('Initials:','xx')
myDlg.addField('SessionNumber:',1)
myDlg.addField('NumberofTrials',40)
myDlg.addField('Time between Display 1 and 2', 1000)
myDlg.addField('Bar length', 100)
myDlg.addField('Bar width', 15)
myDlg.addField('Task (C,O,or E)')
myDlg.show()
if myDlg.OK:
    sessioninfo=myDlg.data
else:
    print('cancelled')
        
def get_display_info():
    """
    gets effective fullscreen resolution from the os and creates a temporary
    window to compute the native hardware resolution and to determine whether
    there is any pixel scaling (i.e., as in MacOS 'Retina' displays)

    returns a tuple consisting of: horizontal resolution, vertical resolution, 
    and pixel scaling factor
    """
    # 1. Get full display resolution (as reported by OS)
    # note: code below assumes either single or twin monitor configuration
    screen = pyglet.canvas.Display().get_default_screen()
    os_width = screen.width
    os_height = screen.height
    # 2. Create fullscreen Psychopy window and get actual hardware resolution
    testwin = visual.Window(monitor='testMonitor',color='black',allowGUI=True,\
        units='pix',size=(os_width,os_height),fullscr=True)
    hard_width,hard_height = testwin.size
    # 3. Compute the ratio of hard_width to os_width to determine whether we're
    #    using a HiDPI display (i.e., one with pixel scaling).
    pixel_scaling = hard_width/os_width
    # 4. Exit the window
    testwin.close()
    return os_width,os_height,pixel_scaling
    

WWIDTH,WHEIGHT,PX_SCALE = get_display_info()
TEXT_SIZE = (WHEIGHT//36)*PX_SCALE      # for instructions (pixels)
TEXT_VPOS = 130*PX_SCALE                # vertical text position (pixels)

#window
mywin = visual.Window([WWIDTH,WHEIGHT], monitor="testMonitor", units="pix",fullscr=True)
#rgb=[-1,-1,-1] makes screen black
random.seed()  #initializes by reading the time


#read values from menu and put in variables
idn=sessioninfo[0]
sessnumber=sessioninfo[1]
ntrials=sessioninfo[2] 
pause=sessioninfo[3]
barlength=sessioninfo[4]*PX_SCALE
barwidth=sessioninfo[5]*PX_SCALE
task=sessioninfo[6]

actual = []
response = []
nbarstostore=[]
taskstore=[]
print(task)

#filename
filename=''
filename='visstm_'
filename+=task[0]+'_'
filename+=idn
filename+='_'+str(sessnumber)
filename+='_'+data.getDateStr()
print(filename)
datetime=data.getDateStr()


#initialize things
displaytime = .25
#(seconds)
#possible x positions of bars...4 choices
barxpos = [-320, -160, 160, 320]
#possible y positions of bars...4 choices
barypos = [-240, -120, 120, 240]
barlocations = []

for x in barxpos:
    for y in barypos:
        barlocations.append([x,y])
colors = ['Red', 'Green', 'Blue', 'Yellow', 'Black', 'White']
orientation = [0, 45, 90, 135]
nbars = [2, 4, 6, 8]


# #write excel file with results
# def write_file(filename):  
#     #write stuff at beginning
#     book = xlwt.Workbook(encoding="utf-8")
#     sheet1 = book.add_sheet("Sheet 1")
#     sheet1.write(0, 0, "Visual Short Term Memory")
#     sheet1.write(1, 0, "Subject:%s"%idn)     
#     sheet1.write(2, 0, "SessionNumber:%d"%sessnumber)
#     sheet1.write(3, 0, "Task:%s" %task)
#     sheet1.write(4, 0, "Inter-display time:%d" %pause)
#     #column labels
#     sheet1.write(11, 0, "Trial")
#     sheet1.write(11, 1, "Task")
#     sheet1.write(11, 2, "N objects")
#     sheet1.write(11, 3, "Actual")
#     sheet1.write(11, 4, "Response")
# #below is info for each trial
#     for i in range (0, ntrials): 
#         sheet1.write(i+12, 0, i+1)          #trial
#         sheet1.write(i+12, 1, taskstore[i])
#         sheet1.write(i+12, 2, nbarstostore[i])
#         sheet1.write(i+12, 3, actual[i])
#         sheet1.write(i+12, 4, response[i])
#         #sheet1.write(i+12, 3, d)    #same or different?       
#     filename+='.xls'
#     book.save(filename)

def displayText(text,vpos=TEXT_VPOS,color='White',size=TEXT_SIZE):
    rendered_text = visual.TextStim(win=mywin,text=text,color=color,\
        pos=(0,vpos),height=size)
    rendered_text.draw() 
    mywin.flip()    

# write excel file with results
def write_file(filename):
    global result
    book = pyxl.Workbook()
    ws = book.active
    ws.title = "Sheet 1"
    # write experiment information
    ws.cell(row=1,column=1).value = "ExperimentName: Visual Short Term Memory"
    ws.cell(row=2,column=1).value = "SubjectID:%s"%idn
    ws.cell(row=3,column=1).value = "SessionNumber:%s"%sessnumber
    ws.cell(row=4,column=1).value = "Task:%s"%task
    ws.cell(row=5,column=1).value = "Inter-display time:%s"%pause
    # write column headers
    headers = ['Trial','Task','SetSize','Actual','Response']
    for col,header in enumerate(headers):
        ws.cell(row=6,column=col+1).value = header
    # write per-trial data
    for r in range(0,len(response)):
        irow = r+7
        ws.cell(irow, 1).value =  r+1
        ws.cell(irow, 2).value = taskstore[r]
        ws.cell(irow, 3).value = nbarstostore[r]
        ws.cell(irow, 4).value = actual[r]
        ws.cell(irow, 5).value = response[r]
    # create directory and/or save file
    if not os.path.exists('data'):
        os.makedirs('data')
    book.save('data/'+filename+'.xlsx')

def choosebars(nbars_trial):                              #randomly select bars for display
    alreadyloc = []
    bars=[0]*nbars_trial  #initialize
    print(nbars_trial)
    for b in range(nbars_trial):
   #     print b
   #     print bars
        barcolor = random.choice(colors)
        barori = random.choice(orientation)
        barloc = random.choice(barlocations)
        while barloc in alreadyloc:
            #check if there's already a bar at this location
            barloc = random.choice(barlocations)
        alreadyloc.append(barloc)
    #    print barlocations
    #    print barloc
        bars[b] = visual.Rect(mywin, width = barwidth, height = barlength)
        bars[b].pos = barloc
        bars[b].fillColor=barcolor
        bars[b].lineColor = barcolor
        bars[b].ori=barori
    z = [b.pos for b in bars]
    r = [b.fillColor for b in bars]
 #   print z
 #   print r
    return bars


        #core.wait(.5)  #uncomment to slow down for debugging
    
#show the display with the bars
def showdisplay(display, bars, diff):
    thistask='s' #default.  will change if display 2 and different
    if display == 2:
        #is second display
        if diff:
            #this display will be different
            d = random.randint(0, len(bars)-1)
            #[print('bar %d color = %s'%(i,el.fillColor)) for i,el in enumerate(bars)]
            if task.lower() == 'e':
                #either...randomly choose color or orientation
                thistask = random.choice(['c','o'])
            else:
                thistask = task.lower()
            if thistask == 'c':
                # Note: the bar colors are stored as numpy arrays (e.g., [1,1,1])
                prevcolor = bars[d].fillColor
                while numpy.all(prevcolor == bars[d].fillColor): #won't be the same color
                    newcolor = random.choice(colors)
                    bars[d].fillColor = newcolor
                    bars[d].lineColor = newcolor
            elif  thistask == 'o':
                prevori = bars[d].ori
                while prevori == bars[d].ori: #won't be the same orientation
                    bars[d].ori = random.choice(orientation)

    for b in range(len(bars)):
        bars[b].draw()
    
    mywin.flip()
    
    core.wait(displaytime)
    return thistask
   #return bars
#displays any text.  Use for response

def blankscreen():
    mywin.flip(clearBuffer=True)
    mywin.flip()
 #   print pause/1000.0
    core.wait(pause/1000.0)

def displaysometext(texttodisplay):
    prompt=visual.TextStim(win=mywin, text=texttodisplay,pos=[0,0],rgb=1,contrast=1)
    prompt.draw()
    mywin.flip()
#get response
def getresponse():
    k = event.waitKeys()
    event.clearEvents()
    b=k[0]
    if b=='escape':  #option to get out
        mywin.close()
        core.quit()
    return b

def waitForValidKeypress(valid_keys):
    key = None
    key = getresponse().upper()
    while key not in valid_keys:
        key = getresponse().upper()
    return key

def showBlockInstructions():
    block_instruction_text = """
    On each trial you’ll see two displays in succession. Each display contains 2, 4, 6, or 8 tilted bars of different colors.
    
    The second display is either the SAME as the first, or it’s DIFFERENT from the first in that one of the items has changed.
    
    After each trial, type S if you think the displays were the same, and type D if you think they were different.
    
    Press SPACEBAR to begin the experiment.
    """
    #winWidth,winHeight = mywin.size

    block_instructions = visual.TextStim(win=mywin,text=block_instruction_text,\
        color='White',pos=(0,0),alignText='center',wrapWidth=0.45*WWIDTH*PX_SCALE,height=TEXT_SIZE) 
    block_instructions.draw()
    mywin.flip()
    # listen for keypress
    keypress = waitForValidKeypress(valid_keys=['SPACE'])
    blankscreen()

showBlockInstructions()
#run all trials
for itrial in range (0,ntrials):
    print("trial %d"%(itrial + 1))
    texttodisplay='Trial ' + repr(itrial+1) + '  Press key' 
    msg = visual.TextStim(win=mywin, text=texttodisplay)
    msg.draw()
    mywin.flip()

    k = ['']
    k = event.waitKeys()
  
    core.wait(1)
    nbarsnow=random.choice(nbars)
#    barstoshow=choosebars(random.choice(nbars)) #    
    barstoshow=choosebars(nbarsnow)
    lettertimer=core.Clock()                        #define a clock to time events in th e trial
    
    diff = bool(random.getrandbits(1))
    #randomly choose if this display will be different or the same
    
    #display displays
    thistask=showdisplay(1, barstoshow, diff)
    blankscreen()
    thistask=showdisplay(2, barstoshow, diff)
    
            
   #Trial over.  Take responses and give feedback         
    
    sdresponse=0
    utargetletterresponse='0'

    displaysometext("same or different?")
    
    while sdresponse not in ['s', 'd']:
        sdresponse=getresponse()
    responsediff = 0
    if sdresponse == 'd':
        sdtext = 'different'
    else:
        sdtext = 'same'
    response.append(sdresponse)
    if diff:
        actualtxt = 'different'
        actual.append('d')
    else:
        actualtxt = 'same'
        actual.append('s')
    feedback='response: ' + sdtext + '   Answer was: ' + actualtxt
#    if responsediff == diff:
#        feedback = 'Correct!'
#    else:
#        feedback = 'Incorrect...'
#    
    displaysometext(feedback)
    core.wait(2)
    
    
    
    #store trial dataFile
    
    nbarstostore.append(nbarsnow)
 #   print 'nbars', nbarstostore
    taskstore.append(str(thistask))
    
    #letter.append(str(target))
    #letter.append(letterstoshow[targetletter])
#    print letter
#    letterseq.append(targetletter+1)
#    print letterseq
#    letterresp.append(targetletterresponse)
#    print letterresp
#    gratingtilt.append(tilt)
#    print gratingtilt
#    gratingseq.append(letterwithgrating+1)
#    print gratingseq
#    gratinglocation.append(int(gloc))
#    print gratinglocation
#    gratingTresponse.append(ugratingtiltresponse)
#    print gratingTresponse
#    gratinglocresponse.append(int(gratinglocationresponse))
#    print gratinglocresponse
#    
   
   
   # if len(event.getKeys())>0: break
   # event.clearEvents()
    
    
   
write_file(filename)
mywin.close()
core.quit()