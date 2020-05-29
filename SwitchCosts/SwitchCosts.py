#Switch Cost
#Jason 8/11/2016: fixed typo in writing results file: == instead of = when writing 'low' or 'even'
# also uResponse instead of userResponse on the same lines
#Christina July 2016
#Based on Monsell, Sumner, and Waters, Memory and Cognition 2003. 31(3) 327 - 342
#Each trial: one number is displayed, either red or green. If the given number is red
#the user must response as to whether it is an odd or even number, if the given numbers
#is green then the user must respond as to whether or not it is higher or lower than 5
#Each experiment is either predictable (changes color and task every 5) or unpredictable
#(color changes at random)

from psychopy import visual, core, gui, data, event, sound
from psychopy.tools.filetools import fromFile, toFile
import time, numpy, random, string, scipy
#import xlwt
import openpyxl as pyxl
import os


global mywin
global grating
global fixation
global blankletter
global centraltext
global xecc,yecc
global savetilt


# set up the menu for choice of conditions
myDlg = gui.Dlg(title="Switch Cost",size=(1, 1))
myDlg.addField('Initials:','xx')
myDlg.addField('SessionNumber:',1)
myDlg.addField('NumberofTrials',30)
myDlg.addField('Intertrial Time ', 2)
myDlg.addField('Type of Task (U or P)')
myDlg.show()
if myDlg.OK:
    sessioninfo=myDlg.data
else:
    print('cancelled')
    
#read values from menu and put in variables
idn=sessioninfo[0] # Sets idn equal to initials
sessnumber=sessioninfo[1] #sets equal to Session nummber
ntrials=sessioninfo[2] #sets n trial equal to number of trials
intertrialTime=sessioninfo[3]#sets equal to intertrial time
tasktype=sessioninfo[4] 


actual = [] #the correct response
response = [] # given response by the user
digitcolor = [] # color of the digit
numValue = [] # value of the number
taskstore = []#the task being performed at a given instant
rTime = [] #reaction times for each experiment
lettPress = [] #the letter the subject pressed


#filename
filename=''
filename='Switchcost'
filename+=idn
filename+='_'+str(sessnumber)
filename+='_'+data.getDateStr()
print(filename)
datetime=data.getDateStr()

#window #*******************************Question Default Screen Color
mywin = visual.Window([800,600], monitor="testMonitor", units="pix")
#rgb=[-1,-1,-1] makes screen black
random.seed()  #initializes by reading the time

global digitVal
global digitColor
global endAnswer



endAnswer = 0
numbers = [ 1, 2 , 3, 4, 6, 7, 8, 9]
colors = ['Red', 'Green'] #two uses, to set the color on the display screen and also to choose the first color of the pred sequence

def blankscreen():
    
    mywin.flip(clearBuffer=True)
    mywin.flip()
 #   print pause/1000.0
    
    
def displaysometext(texttodisplay):
    prompt=visual.TextStim(win=mywin, text=texttodisplay,pos=[0,0],rgb=2,contrast=1)
    prompt.draw()
    mywin.flip()
    

# #write excel file with results
# def write_file(filename):
#     #write stuff at beginning
#     book = xlwt.Workbook(encoding="utf-8")
#     sheet1 = book.add_sheet("Sheet 1")
#     sheet1.write(0, 0, "Switch Cost")
#     sheet1.write(1, 0, "Subject:%s"%idn)     
#     sheet1.write(2, 0, "SessionNumber:%d"%sessnumber)
#     sheet1.write(3, 0, "Task:%s" %tasktype)
#     sheet1.write(4, 0, ":%d" %intertrialTime)
#     #column labels
#     sheet1.write(11, 0, "Trial")
#     sheet1.write(11, 1, "Task")
#     sheet1.write(11, 2, "Actual")
#     sheet1.write(11, 3, "Response")
#     sheet1.write(11, 4, "Reaction Time")
#     sheet1.write(11, 5, "Number Shown")
#     sheet1.write(11, 6, "Letter Pressed")
# #below is info for each trial
#     for i in range (0, ntrials): 
#         sheet1.write(i+12, 0, i+1)          #trial
#         sheet1.write(i+12, 1, taskstore[i])
#         sheet1.write(i+12, 2, actual[i])
#         sheet1.write(i+12, 3, response[i])
#         sheet1.write(i+12, 4, float("%8.3f"%rTime[i])) #THIS IS THE REACTION TIME
#         sheet1.write(i+12, 5, numValue[i])
#         sheet1.write(i+12, 6, lettPress[i])
#     filename+='.xls'
#     book.save(filename)

def write_file(filename):
    global result
    book = pyxl.Workbook()
    ws = book.active
    ws.title = "Sheet 1"
    # write experiment information
    ws.cell(row=1,column=1).value = "ExperimentName: Switch Cost"
    ws.cell(row=2,column=1).value = "SubjectID:%s"%idn
    ws.cell(row=3,column=1).value = "SessionNumber:%s"%sessnumber
    ws.cell(row=4,column=1).value =  "Task:%s" %tasktype
    ws.cell(row=5,column=1).value = ":%d" %intertrialTime
    # write column headers
    headers = ['Trial','Task','Actual','Response','Reaction Time','Number Shown','Letter Pressed']
    for col,header in enumerate(headers):
        ws.cell(row=6,column=col+1).value = header
    # write per-trial data
    for r in range(0,len(response)):
        irow = r+7
        ws.cell(irow, 1).value =  r+1
        ws.cell(irow, 2).value = taskstore[r]
        ws.cell(irow, 3).value = actual[r]
        ws.cell(irow, 4).value = response[r]
        ws.cell(irow, 5).value = float("%2.3f"%rTime[r])
        ws.cell(irow, 6).value = numValue[r]
        ws.cell(irow, 7).value = lettPress[r]
    # create directory and/or save file
    if not os.path.exists('data'):
        os.makedirs('data')
    book.save('data/'+filename+'.xlsx')


def getresponse():
    k = event.waitKeys()
    event.clearEvents()
    b=k[0]
    if b=='escape':  #option to get out
        mywin.close()
        core.quit()
    return b


def responseCheck(chosenColor, digitVal):#Where responses are checked and recorded
    global uResp
    uResp = 0
    userResp = 0
    userResponse = 0
                        
    if chosenColor == 'Green':
        while userResponse not in ['a','l']:
            userResponse=getresponse() 
                                

            currentTime = reactionTime.getTime()
            rTime.append(currentTime)                    
            print(reactionTime)
            reactionTime.reset = 0.0
            
            lettPress.append(userResponse)
            
            if userResponse.lower() =='a': # lOW 
                uResp = 'low'
                response.append(uResp)
                
                
            else:
                uResp = 'high'
                response.append(uResp)
                
                #Something that can be recorded in textfile
                #insert some list or so that can keep tally of how many they get wrong or right for the end 
                # ask eileen if she still wants to tell them what they got wrong/right in the end
                
            if int(digitVal) >= 6: 
                correct = 'high'
                actual.append(correct)
                    
            else:
                correct = 'low'
                actual.append(correct)
            
    if chosenColor == 'Red':                        
        while userResponse not in ['a','l']:
            userResponse=getresponse() 
            
            lettPress.append(userResponse)
            
            currentTime = reactionTime.getTime()
            rTime.append(currentTime)                    
            print(reactionTime)
            reactionTime.reset = 0.0
                    
            if userResponse.lower() == 'a': # lOW 
                uResp = 'odd'
                response.append(uResp)
            else:
                uResp = 'even'
                response.append(uResp)
               
            if int(digitVal) % 2 == 0:
                correct = 'even'
                actual.append(correct)
            else:
                correct = 'odd'
                actual.append(correct)
        
        
def showdisplay(tasktype):
    global chosenColor 
    global x
    global y
    global reactionTime

    
    if tasktype.lower() =='p':
        if chosenColor == 'Green' and y <= 5 :
           
            y = y + 1 #Adds to the counter

            taskstore.append('L/H') 
            digitColor = colors[1]
            digitVal = random.choice(numbers)
                
            print (digitVal)
             
            numValue.append(digitVal)
           
            texttodisplay = str(digitVal)
            msg = visual.TextStim(win=mywin, text=texttodisplay, color = digitColor, height = 72.0)
                      
            msg.draw()
            mywin.flip()
            
            reactionTime = core.Clock()
            
            responseCheck(chosenColor, digitVal)
            
            if y == 5:
                y = 0
                chosenColor = 'Red'
            
            reactionTime = 0.0
            blankscreen()
            core.wait(intertrialTime)

        elif chosenColor == 'Red' and x <= 5: 
            x = x + 1 #Adds to the counter

                
            taskstore.append('O/E') 
            digitVal = random.choice(numbers)
            digitColor = colors[0]
                    
            print(digitVal)
            
            numValue.append(digitVal)
            
 
                    
            texttodisplay = str(digitVal)
            msg = visual.TextStim(win=mywin, text=texttodisplay, color = digitColor, height = 72.0 )
                   
            msg.draw()
            mywin.flip()
            
            reactionTime = core.Clock()

            responseCheck(chosenColor, digitVal) 
                          
            if x == 5:
                x = 0
                chosenColor = 'Green'
            
            reactionTime = 0.0
            blankscreen()
            core.wait(intertrialTime)
    elif tasktype.lower() == 'u':

        digitVal = random.choice(numbers)
        digitColor = random.choice(colors)
        
        print(digitVal)
        
        numValue.append(digitVal)
           
        
        chosenColor = digitColor  
        
        if chosenColor == "Green":
            taskstore.append('L/H')
        elif chosenColor == "Red":
            taskstore.append('O/E')
           
            
        texttodisplay = str(digitVal)
        msg = visual.TextStim(win=mywin, text=texttodisplay, color = digitColor, height = 72.0 )
        
        msg.draw()
        mywin.flip()
            
        reactionTime = core.Clock()
        
        responseCheck(chosenColor, digitVal)

        blankscreen()
        core.wait(intertrialTime)
                    



            
            





#The actual experiment running

        
pInstructions = "If numeral is red: \nPress 'A' if odd.\nPress 'L' if even.\nIf numeral is green:\nPress 'A' if less than 5.\nPress 'L' if greater than 5.\nRespond as fast and as accurately as you can."

print(pInstructions)
 
          
texttodisplay = str(pInstructions)
msg = visual.TextStim(win=mywin, text=texttodisplay, color = 'White', pos = (-100.0,100.0), height = 30)
msg.draw()                                                      #draw (actually, writes to a buffer)
mywin.flip()                                                    #flip = display now
        
k = ['']
k = event.waitKeys()                                        #psychopy function. wait for button press

chosenColor = random.choice(colors)    
x = 0
y = 0

for i in range(0,ntrials):
    
    showdisplay(tasktype)


for i in range(0,ntrials): #KEEPS TALLY OF CORRECT NUMBERS
    if actual[i] == response[i]:
        endAnswer = endAnswer + 1



endTotal = "You got %s out of %s"%(endAnswer, ntrials)
print(endTotal)
  
texttodisplay = str(endTotal)
msg = visual.TextStim(win=mywin, text=texttodisplay, color = 'White')
msg.draw()                                              
mywin.flip()                                                    

write_file(filename)
mywin.close()
core.quit()













