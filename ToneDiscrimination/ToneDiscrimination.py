import psychopy
# set preference for audio sound engine
psychopy.prefs.hardware['audioLib'] = ['PTB', 'sounddevice','pygame']
from psychopy import core, visual, gui, data, event, sound
from psychopy.tools.filetools import fromFile, toFile
import time, numpy, random, scipy
from collections import OrderedDict
#import xlwt
import openpyxl as pyxl
import os

# Number of comparison frequencies & Intertone interval (s)
numFrequencies = 7
intertoneInterval = 1

# Try loading last parameter file, or bring up dialogue box (with default params) and prompt for changes

expInfo = {'Student':'','Standard Frequency (Hz)':1000,'Comparison Increment (Hz)':2,
            'Tone Duration':1,'N trials/comparison':10,
            'Which Tone First':'Standard','Tone Volume (0-1)':0.5,'Session':''}

# Add todays date to the experiment info
expInfo['dateStr']=data.getDateStr()

# Save parameter file (or quit if subject presses cancel)
dlg = gui.DlgFromDict(expInfo,title='Pitch Discrimination - MCS',fixed=['dateStr','Which Tone First'],
        order=['Student','Session'])

# Volume (arbitrary units) range for test
testRange=numpy.linspace(0.1,1.0,10)
# Test tone (1000 Hz, 1 s)
testTone = sound.Sound(1000,expInfo['Tone Duration'])

# Find maximum and minimum comparison frequencies, given the Standard and the Comparison Increment
maxComparison = expInfo['Standard Frequency (Hz)']+expInfo['Comparison Increment (Hz)']*(numFrequencies//2)
minComparison = expInfo['Standard Frequency (Hz)']-expInfo['Comparison Increment (Hz)']*(numFrequencies//2)
# Create array of comparison frequencies to be used for experiment
comparisonFreqs = numpy.linspace(minComparison,maxComparison,numFrequencies)
# Build array of all comparison tones to be tested (size = numFrequencies*N trials/comparison). Then randomize the tone order in the array
comparisonFreq = []
for i in range(expInfo['N trials/comparison']):
    comparisonFreq=numpy.append(comparisonFreq,comparisonFreqs)
random.shuffle(comparisonFreq)

# Create data file to write data to
fileName='PitchDiscrimination_'+expInfo['Student']+'_'+expInfo['Session']+'_'+expInfo['dateStr']
#dataFile=open(fileName+'.xls','w')
# Create column headers for data file
#dataFile.write("First Tone (S or C), Standard (Hz), Comparison (Hz), Duration (s), Loudness, Response (H or L) \n")

# Bring up experiment window
win=visual.Window([800,600],allowGUI=True,monitor='testMonitor',units='deg')

# define function getKeysWithExit to allow exiting at any keypress
def waitKeysWithExit(keyList=None):
    keys = event.waitKeys(keyList=keyList)
    for current_key in keys:
        if current_key in ['q','escape']:
            win.close()
            core.quit()
    return keys

# Set up structure for Excel workbook
book = pyxl.Workbook()
ws = book.active
ws.title = "Sheet 1"
# write experiment information
ws.cell(row=1,column=1).value = "ExperimentName: Tone Discrimination"
ws.cell(row=2,column=1).value = "SubjectID:%s"%expInfo['Student']
ws.cell(row=3,column=1).value = "SessionNumber:%s"%expInfo['Session']
ws.cell(row=4,column=1).value = "Loudness:%s"%expInfo['Tone Volume (0-1)']
ws.cell(row=5,column=1).value = "Which first %s"%expInfo['Which Tone First']
ws.cell(row=6,column=1).value = "Standard:%d"%expInfo['Standard Frequency (Hz)']
ws.cell(row=7,column=1).value = "Duration(s)%f"%expInfo['Tone Duration']
# write column headers
headers = ['Trial','1stTone','Standard(Hz)','Comparison(Hz)','Response']
for col,header in enumerate(headers):
    ws.cell(row=9,column=col+1).value = header

# Initialize clock
globalClock=core.Clock()
trialClock=core.Clock()

# Message informing subject to adjust volume (must type 'volume' to continue)
messageVolumeAdjust=visual.TextStim(win,pos=[0,0],text="This exercise will be using tones as stimuli. \n\n \
Be sure to set the volume of your speakers to a low level!!! \n \
If it is too low you can always increase slightly and re-start. \n \
This is particularly important for those using headphones. \n\n \
Type 'v' when you've adjusted your speaker volume to a low level.")

# Make subject retype 'volume'
messageConfirmation=visual.TextStim(win,pos=[0,0],text="Retype 'v' to confirm your speakers have been adjusted.")

# Message to display while testing speaker volume
messageTest=visual.TextStim(win,pos=[0,0],text="Testing speaker volume. \n\n \
If ANY tones are too soft or loud, adjust speaker volume and restart (press q or esc). \n\n \
Otherwise, wait for the tones to finish and then press any key to continue.")

# Message introducing subject to experiment (\n adds line space)
messageIntro=visual.TextStim(win,pos=[0,0],text="Two tones will be presented, a Standard Tone and a Comparison Tone. \n\n \
The %s tone will be presented first. \n\n \
A total of %i trials will be run. \n\n \
Do not adjust the volume of the audio output during the experiment. \n\n \
(press any key to continue)" %(expInfo['Which Tone First'],(expInfo['N trials/comparison']*numFrequencies)))

# Message giving instructions for running experiment
messageInstructions=visual.TextStim(win,pos=[0,0],text="Use the arrow keys to indicate the pitch of the Comparison relative to the Standard. \n\n \
Press up arrow if the Comparison has a higher pitch. \n\n \
Press down arrow if the Comparison has a lower pitch. \n\n \
Guess if you're not sure. \n\n \
(press any key to continue)")

# Message to be used when prompting subject for response. If statement ensures that subject will always be reminded which
# tone the Comparison was (first or second)
if expInfo['Which Tone First']=='Standard':
    messageResponse=visual.TextStim(win,pos=[0,+3],text="Comparison (second tone) higher pitch -> press up arrow. \n\n \
    Comparison (second tone) lower pitch -> press down arrow.")
elif expInfo['Which tone First']=='Comparison':
    messageResponse=visual.TextStim(win,pos=[0,+3],text="Comparison (first tone) higher pitch -> press up arrow. \n\n \
    Comparison (first tone) lower pitch -> press down arrow.")

# Make standard tone and set its volume
standardTone = sound.Sound(expInfo['Standard Frequency (Hz)'],expInfo['Tone Duration'])
standardTone.setVolume(expInfo['Tone Volume (0-1)'])

# Show volume adjustment message and wait for user to type 'v'
messageVolumeAdjust.draw()
win.flip()
event.waitKeys(keyList='v')

# Show confirmation message and wait for user to type 'v'
messageConfirmation.draw()
win.flip()
event.waitKeys(keyList='v')

# Show volume test message
messageTest.draw()
win.flip()
# Play test tones at each volume
for thisVolume in testRange:
    testTone.setVolume(thisVolume)
    testTone.play()
    core.wait(expInfo['Tone Duration'])
    core.wait(intertoneInterval)
allKeys = event.waitKeys()
# Quit experiment if subject presses q or esc
for thisKey in allKeys:
    if thisKey in ['q','escape']:
        win.close()
        core.quit()

# Show introductory message and wait for key press
messageIntro.draw()
win.flip()
event.waitKeys()

# Show instructions and wait for key press
messageInstructions.draw()
win.flip()
event.waitKeys()

# Set current trial to 1
currentTrial = 1

for thisFreq in comparisonFreq:
    # Create Comparison tone and set volume for this trial
    comparisonTone = sound.Sound(thisFreq,expInfo['Tone Duration'])
    comparisonTone.setVolume(expInfo['Tone Volume (0-1)'])
    # Present current trial # and wait for key press to initiate trial
    messageTrial=visual.TextStim(win,pos=[0,+3],text="Trial %i (press any key to continue)" %currentTrial)
    messageTrial.draw()
    win.flip()
    event.waitKeys()
    # Flip to grey screen after key press so subject knows trial is initiated
    win.flip()
    
    # If Standard first, wait for intertone interval, play Standard, pause for intertone interval and then play Comparison
    core.wait(intertoneInterval)
    if expInfo['Which Tone First']=='Standard':
        standardTone.play()
        core.wait(expInfo['Tone Duration'])
        core.wait(intertoneInterval)
        comparisonTone.play()
        core.wait(expInfo['Tone Duration'])
    # If Comparison first, wait for intertone interval, play Comparison, pause for intertone interval, then play Standard
    elif expInfo['Which Tone First']=='Comparison':
        comparisonTone.play()
        core.wait(expInfo['Tone Duration'])
        core.wait(intertoneInterval)
        standardTone.play()
        core.wait(expInfo['Tone Duration'])

    # Prompt participant for response
    messageResponse.draw()
    win.flip()
    
    # Wait for response
    thisResp=None
    while thisResp==None:
        allKeys=event.waitKeys()
        for thisKey in allKeys:
            # if subject presses down key, response is L (lower)
            if thisKey=='down':
                thisResp='L'
            # If subject presses up key, response is H (higher)
            elif thisKey=='up':
                thisResp='H'
            # If subject press Q or ESC, end experiment
            elif thisKey in ['q','escape']:
                win.close()
                core.quit()
        event.clearEvents()

    # Write data from this trial to data file
    irow = currentTrial+10
    ws.cell(irow, 1).value =  currentTrial
    ws.cell(irow, 2).value = "%s"%expInfo['Which Tone First']
    ws.cell(irow, 3).value = expInfo['Standard Frequency (Hz)']
    ws.cell(irow, 4).value = thisFreq
    ws.cell(irow, 5).value = thisResp
    currentTrial = currentTrial+1

    

   

# create data directory and save file
if not os.path.exists('data'):
    os.makedirs('data')
book.save('data/'+fileName+'.xlsx')


# Close window and quit experiment
win.close()
core.quit()