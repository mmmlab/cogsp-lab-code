#Attentional Blink
#Asante April 2017; revised and refactored extensively by Melchi -- June 2020
#The Attentional Blink Experiment was designed to present a series of stimuli and have the user find two images. The first will 
#always come before the second, however due to the Attentional Blink that happens in memory after finding the first image, keeps
#the user from consciously finding the 2nd image.
import psychopy
# set preference for audio sound engine
psychopy.prefs.hardware['audioLib'] = ['PTB', 'pyo','pygame']
from psychopy import visual, core, gui, data, event, sound
from psychopy.tools.filetools import fromFile, toFile
import time, numpy, string, scipy
from numpy import random
import openpyxl as pyxl
import pyglet
import os

################################################################################
## Utility Function to Compute Screen Resolution and Pixel Scaling
def get_display_info():
    """
    gets effective fullscreen resolution from the os and creates a temporary
    window to compute the native hardware resolution and to determine whether
    there is any pixel scaling (i.e., as in MacOS 'Retina' displays)

    returns a tuple consisting of: horizontal resolution, vertical resolution, 
    and pixel scaling factor
    """
    # 1. Get full display resolution (as reported by OS)
    # note: code below assumes either single or twin monitor configuration
    screen = pyglet.canvas.Display().get_default_screen()
    os_width = screen.width
    os_height = screen.height
    # 2. Create fullscreen Psychopy window and get actual hardware resolution
    testwin = visual.Window(monitor='testMonitor',color='black',allowGUI=True,\
        units='pix',size=(os_width,os_height),fullscr=True)
    hard_width,hard_height = testwin.size
    # 3. Compute the ratio of hard_width to os_width to determine whether we're
    #    using a HiDPI display (i.e., one with pixel scaling).
    pixel_scaling = hard_width/os_width
    # 4. Exit the window
    testwin.close()
    return os_width,os_height,pixel_scaling
################################################################################

## Global Constants and Variables
global mywin

WWIDTH,WHEIGHT,PX_SCALE = get_display_info()
LETTERS = list('ABCDEFGHIJKLMNOPQRSTUVWYZ')
FIXATION_SIZE = 80*PX_SCALE # pixels
LETTER_SIZE = 100*PX_SCALE  # for stimuli (pixels)
TEXT_SIZE = 30*PX_SCALE     # for instructions (pixels)
TEXT_VPOS = 130*PX_SCALE    # vertical text position (pixels)
STIMULUS_DURATION = 0.015   # seconds
ISI_DURATION = 0.085        # seconds

beforeWhite = []      #number of letters prior to white letter
targetLetter = []     #letter that is in white
containSet = []       #whether there was an X or not
afterWhite = []       #how many letters after the white one the X came up
respX = []            # if they responded whether there was an x or not
respTarget = []       #Response as to what the target letter was

## Functions
def getResponse():
    k = event.waitKeys()
    event.clearEvents()
    b=k[0]
    if b=='escape':  #option to get out
        mywin.close()
        core.quit()
    return b

def waitForValidKeypress(valid_keys):
    key = None
    key = getResponse().upper()
    while key not in valid_keys:
        key = getResponse().upper()
    return key

def blankScreen(): 
    mywin.flip(clearBuffer=True)
    mywin.flip()

def displayText(text,vpos=TEXT_VPOS,color='White',size=TEXT_SIZE):
    rendered_text = visual.TextStim(win=mywin,text=text,color=color,\
        pos=(0,vpos),height=size)
    rendered_text.draw() 
    mywin.flip()    

def write_file(filename):
    book = pyxl.Workbook()
    ws = book.active
    ws.title = "Sheet 1"
    # write experiment information
    ws.cell(row=1,column=1).value = "ExperimentName: Attentional Blink"
    ws.cell(row=2,column=1).value = "SubjectID:%s"%subid
    ws.cell(row=3,column=1).value = "SessionNumber:%s"%session
    # write column headers
    headers = ['Trial','TargetPosition','TargetLetter','TargetResponse',\
        'X_Position','X_Present','X_Response']
    ws.append(headers)
    # write per-trial data
    for i in range(0,ntrials):
        row_data = [i+1,beforeWhite[i],targetLetter[i],respTarget[i],\
            afterWhite[i],containSet[i],respX[i]]
        ws.append(row_data)
    # create data directory and save data file
    if not os.path.exists('data'):
        os.makedirs('data')
    book.save('data/'+filename+'.xlsx')

def createRSVPSequence(include_x):
    ## TODO:
    # This is a bit bizarre, but I'm copying what was done in the original code,
    # which varied the length of the RSVP sequence depending on the position of
    # the white letter.
    # We should review the methods of the relevant readings to check whether 
    # this is necessary and/or desireable.
    LENGTH_AFTER_TARGET = 8
    target_index = random.randint(7,16)
    sequence_length = target_index + LENGTH_AFTER_TARGET
    # create the sequence
    rsvp_sequence = random.choice(LETTERS,sequence_length,replace=True)
    target_value = rsvp_sequence[target_index]
    if(include_x):
        x_offset = random.randint(1,LENGTH_AFTER_TARGET)
        x_index = x_offset+target_index
        rsvp_sequence[x_index] = 'X'
    else:
        x_offset = None
    return rsvp_sequence,target_value,target_index,x_offset

def drawLetter(letter,fontcolor):
    letter_stim = visual.TextStim(win=mywin, text=letter, color=fontcolor,\
        pos=(0,0),height = LETTER_SIZE) 
    letter_stim.draw()

def displayStimulusSequence(sequence,target_index):
    for i,letter in enumerate(sequence):
        if(i==target_index):
            drawLetter(letter,'White')
        else:
            drawLetter(letter,'Black')
        mywin.flip()
        core.wait(STIMULUS_DURATION)
        blankScreen()
        core.wait(ISI_DURATION)

def runTrial(trialType):
    # Present fixation/trial start screen
    press_key_text = "Press SPACEBAR to start the trial."
    press_key = visual.TextStim(win=mywin, text=press_key_text, color = 'Black',
        pos = (0,TEXT_VPOS),height=TEXT_SIZE)  
    crossImg = visual.TextStim(win=mywin, text='+', color = 'Black',
        pos = (0,0),height=FIXATION_SIZE) 
    press_key.draw() 
    crossImg.draw()                                                      
    mywin.flip()
    # listen for keypress
    keypress = waitForValidKeypress(valid_keys=['SPACE'])
    blankScreen()
    core.wait(.5)
    ############################################################################    
    ## Generate and Present Stimulus
    # generate trial stimuli
    stim_sequence,target,target_idx,x_offset = createRSVPSequence(trialType)
    # display the stimulus sequence
    displayStimulusSequence(stim_sequence,target_idx)
    ############################################################################    
    ## Target Letter Identification Query       
    letter_query_text = "Which letter was white? Press the corresponding letter on the keyboard"
    displayText(letter_query_text)
   
    # listen for keypress response (any letter is a valid response)
    valid_responses = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    target_response = waitForValidKeypress(valid_keys=valid_responses)
    blankScreen()    
    ############################################################################
    ## X Presence Query
    x_query_text = "Was there an x in this trial? Press 'n' for no and 'y' for yes"
    displayText(x_query_text,vpos=150)
    # listen for keypress response
    keypress = waitForValidKeypress(valid_keys='YN')
    x_response = 1 if keypress=='Y' else 0
    blankScreen()
    ############################################################################
    ## Store trial parameters and responses
    afterWhite.append(x_offset)
    targetLetter.append(target)
    beforeWhite.append(target_idx)
    containSet.append(trialType)
    # save the observer's responses
    respX.append(str(x_response)) 
    respTarget.append(target_response)

def runBlock():
    block_instruction_text = (
    "A series of letters will be shown. You have two tasks:\n\n"
    "(1) Within the series is a white letter. Find and identify the letter.\n\n"
    "(2) After this letter appears another series of letters will be shown, and"
    " you will need to determine whether this second sequence contains an X.\n\n"
    "Following the complete presentation of the letter sequence will be asked"
    " to report which letter was in white and whether there was an X.\n\n"
    "Press SPACEBAR to start the experiment."
    )
    #winWidth,winHeight = mywin.size

    block_instructions = visual.TextStim(win=mywin,text=block_instruction_text,\
        color='White',pos=(0,0),alignText='left',wrapWidth=800*PX_SCALE,height=TEXT_SIZE) 
    block_instructions.draw()
    mywin.flip()
    # listen for keypress
    keypress = waitForValidKeypress(valid_keys=['SPACE'])
    blankScreen()
    ############################################################################
    ## Trial Loop
    for i in range(ntrials):
        trialType = random.randint(0,2)
        runTrial(trialType)
    # exit screen
    exit_text = "Press any key to exit the program"
    displayText(exit_text,color='Black',vpos=90)
    keypress = getResponse()    

################################################################################
##                  Experiment Script Starts Here                             ##
################################################################################
## Setup and Run GUI Menu 
myDlg = gui.Dlg(title="Attentional Blink",size=(1, 1))
myDlg.addField('Initials:','xx')
myDlg.addField('SessionNumber:',1)
myDlg.addField('NumberofTrials',15)
myDlg.show()
if myDlg.OK:
    sessioninfo=myDlg.data
else:
    print('cancelled') #Changed this from print 'cancelled' to print('cancelled')
## Extract GUI Menu Values
subid=sessioninfo[0] # Sets subid equal to initials
session=sessioninfo[1] #sets equal to Session nummber
ntrials=sessioninfo[2] #sets n trial equal to number of trials
## Create Display Window
mywin = visual.Window(size=(WWIDTH,WHEIGHT),fullscr=True,screen=0,\
    monitor="testMonitor",units="pix",color = (0,0,0))
# initialize random seed based on current time
random.seed()
## Run a Block of Trials, Save Data, and Exit Program
runBlock()
# construct output file name
filename='AttentionalBlink_%s_%03d_%s'%(subid,session,data.getDateStr())
# save experiment data
write_file(filename)  
mywin.close()
core.quit()