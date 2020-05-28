#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
This experiment was created using PsychoPy2 Experiment Builder (v1.82.01), November 13, 2015, at 16:55
If you publish work using this script please cite the relevant PsychoPy publications
  Peirce, JW (2007) PsychoPy - Psychophysics software in Python. Journal of Neuroscience Methods, 162(1-2), 8-13.
  Peirce, JW (2009) Generating stimuli for neuroscience using PsychoPy. Frontiers in Neuroinformatics, 2:10. doi: 10.3389/neuro.11.010.2008
"""

from __future__ import division  # so that 1/3=0.333 instead of 1/3=0
from psychopy import visual, core, data, event, logging, sound, gui
from psychopy.constants import *  # things like STARTED, FINISHED
import numpy as np  # whole numpy lib is available, prepend 'np.'
from numpy import sin, cos, tan, log, log10, pi, average, sqrt, std, deg2rad, rad2deg, linspace, asarray
from numpy.random import random, randint, normal, shuffle
import os  # handy system and path functions

# Ensure that relative paths start from the same directory as this script
_thisDir = os.path.dirname(os.path.abspath(__file__))
os.chdir(_thisDir)

# Store info about the experiment session
expName = 'Categorization'  # from the Builder filename that created this script
expInfo = {u'Subject_ID': u'', u'Session_Number': u'', u'Boolean_Complexity': u''}
dlg = gui.DlgFromDict(dictionary=expInfo, title=expName)
if dlg.OK == False: core.quit()  # user pressed cancel
expInfo['date'] = data.getDateStr()  # add a simple timestamp
expInfo['expName'] = expName

# Data file name stem = absolute path + name; later add .psyexp, .csv, .log, etc
filename = _thisDir + os.sep + 'data/%s_%s_%s' %(expInfo['Subject_ID'], expName, expInfo['date'])

# An ExperimentHandler isn't essential but helps with data saving
thisExp = data.ExperimentHandler(name=expName, version='',
    extraInfo=expInfo, runtimeInfo=None,
    originPath=None,
    savePickle=True, saveWideText=True,
    dataFileName=filename)
#save a log file for detail verbose info
logFile = logging.LogFile(filename+'.log', level=logging.DATA)
logging.console.setLevel(logging.WARNING)  # this outputs to the screen, not a file

endExpNow = False  # flag for 'escape' or other condition => quit the exp

# Start Code - component code to be run before the window creation

# Setup the Window
win = visual.Window(size=(1366, 768), fullscr=True, screen=0, allowGUI=True, allowStencil=False,
    monitor='testMonitor', color=[0,0,0], colorSpace='rgb',
    blendMode='avg', useFBO=True,
    )
# store frame rate of monitor if we can measure it successfully
expInfo['frameRate']=win.getActualFrameRate()
if expInfo['frameRate']!=None:
    frameDur = 1.0/round(expInfo['frameRate'])
else:
    frameDur = 1.0/60.0 # couldn't get a reliable measure so guess

# Initialize components for Routine "instructPage"
instructPageClock = core.Clock()
topLeft = visual.TextStim(win=win, ori=0, name='topLeft',
    text='Exp Name : Categorization',    font='Arial',
    pos=[-0.8, 0.8], height=0.05, wrapWidth=None,
    color='white', colorSpace='rgb', opacity=1,
    depth=0.0)
topRight = visual.TextStim(win=win, ori=0, name='topRight',
    text='More Information:\n1) Info Line 1\n2) Info Line 2',    font='Arial',
    pos=[0.8, 0.8], height=0.07, wrapWidth=None,
    color='white', colorSpace='rgb', opacity=1,
    depth=-1.0)
instructions = visual.TextStim(win=win, ori=0, name='instructions',
    text=u'Instructions\n\n(1) Study Phase (10 seconds)\nStudy the objects. \nSome are labeled \u201cIn Category,\u201d and others labeled \u201cNot In Category.\u201d\n\n(2) Test Phase\nOne object is presented at a time.\nPress \u201cz\u201d if it is In Category.\nPress \u201cm\u201d if it is Not In Category.\n\nThe Study and Test Phases will repeat for a different set of objects.',    font='Arial',
    pos=[0, 0], height=0.06, wrapWidth=None,
    color='white', colorSpace='rgb', opacity=1,
    depth=-2.0)
getContinue = visual.TextStim(win=win, ori=0, name='getContinue',
    text='Press Spacebar to start.',    font='Arial',
    pos=[0, -0.8], height=0.09, wrapWidth=None,
    color='white', colorSpace='rgb', opacity=1,
    depth=-3.0)
# import the gui package in PsychoPy

from psychopy import gui

# Initialize components for Routine "trial"
trialClock = core.Clock()
#--import the required packages.
from openpyxl import load_workbook
import random

#--trialOrder will help track the order in which 
#--the categorization problems are chosen.
dirty = 1

#--The follwing lines of code generate the different (shape,color,number) combinations for the test cases.
iniTrials = range(1, 12, 2) #--total number of samples possible
trialOrder = random.sample(iniTrials, 3) #--Selects 3 random test cases from the total list.


exttrialOrder = [x + 1 for x in trialOrder]#--the enfored test cases based on the test cases selected above.
[trialOrder.append(x) for x in exttrialOrder]#--this list contains all the test cases that will be used.
shuffle(trialOrder) #this makes the order of test cases random

#print trialOrder
#--The Code maintains every element as a object of type Stimulus.
#--Each object will have a distinct combination of Shape, Color, Number and parity.
#--Every repeat will have a different combination of objects with random and distinct property assignment.
class Stimulus:
    #--x,y,z will have random assignment of shape, color and number.
    #--the random value assignment can be interpreted from the value in assign
    #--inCat gives the category of the stimuli object
    #--par gives the parity value
    def __init__(self, x, y, z, inCat, par, assign):
        if assign <= 2: #-- x=shape y=number z=color
            self.s = x
            self.n = y
            self.c = z
            self.cat = inCat
            self.p = par
        elif assign <= 4: #-- x=shape y=color z=number
            self.s = x
            self.n = z
            self.c = y
            self.cat = inCat
            self.p = par
        elif assign <= 6: #-- x=number y=shape z=color
            self.s = y
            self.n = x
            self.c = z
            self.cat = inCat
            self.p = par
        elif assign <= 8: #-- x=number y=color z=shape
            self.s = z
            self.n = x
            self.c = y
            self.cat = inCat
            self.p = par
        elif assign <= 10: #-- x=color y=number z=shape
            self.s = z
            self.n = y
            self.c = x
            self.cat = inCat
            self.p = par
        else: #-- x=color y=shape z=number
            self.s = y
            self.n = z
            self.c = x
            self.cat = inCat
            self.p = par


# random 8 "trial" numbers among 16(=2^4) possible combinations of shape, number, color, incat value
# to avoid repeating the test stimuli once tested. --sk

# Make 6 random numbers among 16
testTypeNo = random.sample(range(1,16,1), 6 )
shuffle(testTypeNo)
print testTypeNo

shapeList      = ['square', 'triangle']
numberList     = ['2','1']
stcolorList    = ['black', 'white']
inCatList      = ['0','1']










topTitle = visual.Rect(win=win, name='topTitle',
    width=[0.4, 0.1][0], height=[0.4, 0.1][1],
    ori=0, pos=[0, 0.55],
    lineWidth=4, lineColor=[-1,-1,-1], lineColorSpace='rgb',
    fillColor=[1,1,1], fillColorSpace='rgb',
    opacity=1,depth=-1.0, 
interpolate=True)
inCatText = visual.TextStim(win=win, ori=0, name='inCatText',
    text='In Category',    font='Arial',
    pos=[0, 0.55], height=0.08, wrapWidth=None,
    color='black', colorSpace='rgb', opacity=1,
    depth=-2.0)
bottomTitle = visual.Rect(win=win, name='bottomTitle',
    width=[0.4, 0.1][0], height=[0.4, 0.1][1],
    ori=0, pos=[0, -0.55],
    lineWidth=4, lineColor=[-1,-1,-1], lineColorSpace='rgb',
    fillColor=[1,1,1], fillColorSpace='rgb',
    opacity=1,depth=-3.0, 
interpolate=True)
notinCatText = visual.TextStim(win=win, ori=0, name='notinCatText',
    text='Not In Category',    font='Arial',
    pos=[0, -0.54], height=0.08, wrapWidth=None,
    color='Black', colorSpace='rgb', opacity=1,
    depth=-4.0)
trialmsg = visual.TextStim(win=win, ori=0, name='trialmsg',
    text='Test Number :',    font='Arial',
    pos=[-0.8, 0.85], height=0.05, wrapWidth=None,
    color='white', colorSpace='rgb', opacity=1,
    depth=-5.0)
trialNum = visual.TextStim(win=win, ori=0, name='trialNum',
    text='default text',    font='Arial',
    pos=[-0.8, 0.8], height=0.05, wrapWidth=None,
    color='white', colorSpace='rgb', opacity=1,
    depth=-6.0)
ISI = core.StaticPeriod(win=win, screenHz=expInfo['frameRate'], name='ISI')
categories = visual.Rect(win=win, name='categories',
    width=[1, 1][0], height=[1, 1][1],
    ori=0, pos=[0, 0],
    lineWidth=4, lineColor=[-1,-1,-1], lineColorSpace='rgb',
    fillColor=[1,1,1], fillColorSpace='rgb',
    opacity=1,depth=-8.0, 
interpolate=True)
categorySeparator = visual.Line(win=win, name='categorySeparator',
    start=(-[1, 1][0]/2.0, 0), end=(+[1, 1][0]/2.0, 0),
    ori=0, pos=[0, 0],
    lineWidth=4, lineColor=[-1,-1,-1], lineColorSpace='rgb',
    fillColor=[-1,-1,-1], fillColorSpace='rgb',
    opacity=1,depth=-9.0, 
interpolate=True)

# Initialize components for Routine "trialtest"
trialtestClock = core.Clock()

trialTestMsg = visual.TextStim(win=win, ori=0, name='trialTestMsg',
    text='Test Number :',    font='Arial',
    pos=[-0.8, 0.85], height=0.05, wrapWidth=None,
    color='white', colorSpace='rgb', opacity=1,
    depth=-1.0)
trialTestNum = visual.TextStim(win=win, ori=0, name='trialTestNum',
    text='default text',    font='Arial',
    pos=[-0.8, 0.8], height=0.05, wrapWidth=None,
    color='white', colorSpace='rgb', opacity=1,
    depth=-2.0)
oval = visual.Polygon(win=win, name='oval',
    edges = 100, size=[0.3, 0.2],
    ori=0, pos=[0, 0.3],
    lineWidth=3, lineColor=[-1,-1,-1], lineColorSpace='rgb',
    fillColor=[1,1,1], fillColorSpace='rgb',
    opacity=1,depth=-3.0, 
interpolate=True)
square1 = visual.Rect(win=win, name='square1',
    width=[0.025, 0.05][0], height=[0.025, 0.05][1],
    ori=0, pos=[-0.05, 0.3],
    lineWidth=2, lineColor=[-1,-1,-1], lineColorSpace='rgb',
    fillColor=1.0, fillColorSpace='rgb',
    opacity=1.0,depth=-4.0, 
interpolate=True)
square2 = visual.Rect(win=win, name='square2',
    width=[0.025, 0.05][0], height=[0.025, 0.05][1],
    ori=0, pos=[0, 0.3],
    lineWidth=2, lineColor=[-1,-1,-1], lineColorSpace='rgb',
    fillColor=1.0, fillColorSpace='rgb',
    opacity=1,depth=-5.0, 
interpolate=True)
triangle1 = visual.ShapeStim(win=win, name='triangle1',
    vertices = [[-[0.025, 0.05][0]/2.0,-[0.025, 0.05][1]/2.0], [+[0.025, 0.05][0]/2.0,-[0.025, 0.05][1]/2.0], [0,[0.025, 0.05][1]/2.0]],
    ori=0, pos=[-0.05, 0.3],
    lineWidth=2, lineColor=[-1,-1,-1], lineColorSpace='rgb',
    fillColor=1.0, fillColorSpace='rgb',
    opacity=1.0,depth=-6.0, 
interpolate=True)
triangle2 = visual.ShapeStim(win=win, name='triangle2',
    vertices = [[-[0.025, 0.05][0]/2.0,-[0.025, 0.05][1]/2.0], [+[0.025, 0.05][0]/2.0,-[0.025, 0.05][1]/2.0], [0,[0.025, 0.05][1]/2.0]],
    ori=0, pos=[0, 0.3],
    lineWidth=2, lineColor=[-1,-1,-1], lineColorSpace='rgb',
    fillColor=1.0, fillColorSpace='rgb',
    opacity=1,depth=-7.0, 
interpolate=True)
inputmsg = visual.TextStim(win=win, ori=0, name='inputmsg',
    text='Press "z" to answer In Category.\nPress "m" to answer Not In Cateogry.',    font='Arial',
    pos=[0, -0.6], height=0.1, wrapWidth=None,
    color='white', colorSpace='rgb', opacity=1,
    depth=-8.0)

# Initialize components for Routine "blankSpace"
blankSpaceClock = core.Clock()
ISI_2 = core.StaticPeriod(win=win, screenHz=expInfo['frameRate'], name='ISI_2')

# Initialize components for Routine "nextTrial"
nextTrialClock = core.Clock()
contMsg = visual.TextStim(win=win, ori=0, name='contMsg',
    text='Press any key to go on to the next set.',    font='Arial',
    pos=[0, 0], height=0.06, wrapWidth=None,
    color='White', colorSpace='rgb', opacity=1,
    depth=0.0)
exitMsg = visual.TextStim(win=win, ori=0, name='exitMsg',
    text='Thank you !!\nYou have completed the Experiment.\nPress any key to exit.\n',    font='Arial',
    pos=[0, 0], height=0.07, wrapWidth=None,
    color='White', colorSpace='rgb', opacity=1,
    depth=-1.0)

# Create some handy timers
globalClock = core.Clock()  # to track the time since experiment started
routineTimer = core.CountdownTimer()  # to track time remaining of each (non-slip) routine 

#------Prepare to start Routine "instructPage"-------
t = 0
instructPageClock.reset()  # clock 
frameN = -1
# update component parameters for each repeat
contInput = event.BuilderKeyResponse()  # create an object of type KeyResponse
contInput.status = NOT_STARTED

# keep track of which components have finished
instructPageComponents = []
instructPageComponents.append(topLeft)
instructPageComponents.append(topRight)
instructPageComponents.append(instructions)
instructPageComponents.append(getContinue)
instructPageComponents.append(contInput)
for thisComponent in instructPageComponents:
    if hasattr(thisComponent, 'status'):
        thisComponent.status = NOT_STARTED

#-------Start Routine "instructPage"-------
continueRoutine = True
while continueRoutine:
    # get current time
    t = instructPageClock.getTime()
    frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
    # update/draw components on each frame
    
    # *topLeft* updates
    if t >= 0.0 and topLeft.status == NOT_STARTED:
        # keep track of start time/frame for later
        topLeft.tStart = t  # underestimates by a little under one frame
        topLeft.frameNStart = frameN  # exact frame index
        topLeft.setAutoDraw(True)
    
    # *topRight* updates
    if t >= 0.0 and topRight.status == NOT_STARTED:
        # keep track of start time/frame for later
        topRight.tStart = t  # underestimates by a little under one frame
        topRight.frameNStart = frameN  # exact frame index
        topRight.setAutoDraw(True)
    if topRight.status == STARTED and t >= (0.0 + (0.0-win.monitorFramePeriod*0.75)): #most of one frame period left
        topRight.setAutoDraw(False)
    
    # *instructions* updates
    if t >= 0.0 and instructions.status == NOT_STARTED:
        # keep track of start time/frame for later
        instructions.tStart = t  # underestimates by a little under one frame
        instructions.frameNStart = frameN  # exact frame index
        instructions.setAutoDraw(True)
    
    # *getContinue* updates
    if t >= 0.0 and getContinue.status == NOT_STARTED:
        # keep track of start time/frame for later
        getContinue.tStart = t  # underestimates by a little under one frame
        getContinue.frameNStart = frameN  # exact frame index
        getContinue.setAutoDraw(True)
    
    # *contInput* updates
    if t >= 0.0 and contInput.status == NOT_STARTED:
        # keep track of start time/frame for later
        contInput.tStart = t  # underestimates by a little under one frame
        contInput.frameNStart = frameN  # exact frame index
        contInput.status = STARTED
        # keyboard checking is just starting
        contInput.clock.reset()  # now t=0
        event.clearEvents(eventType='keyboard')
    if contInput.status == STARTED:
        theseKeys = event.getKeys(keyList=['space'])
        
        # check for quit:
        if "escape" in theseKeys:
            endExpNow = True
        if len(theseKeys) > 0:  # at least one key was pressed
            contInput.keys = theseKeys[-1]  # just the last key pressed
            contInput.rt = contInput.clock.getTime()
            # a response ends the routine
            continueRoutine = False
    
    
    # check if all components have finished
    if not continueRoutine:  # a component has requested a forced-end of Routine
        break
    continueRoutine = False  # will revert to True if at least one component still running
    for thisComponent in instructPageComponents:
        if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
            continueRoutine = True
            break  # at least one component has not yet finished
    
    # check for quit (the Esc key)
    if endExpNow or event.getKeys(keyList=["escape"]):
        core.quit()
    
    # refresh the screen
    if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
        win.flip()

#-------Ending Routine "instructPage"-------
for thisComponent in instructPageComponents:
    if hasattr(thisComponent, "setAutoDraw"):
        thisComponent.setAutoDraw(False)
# check responses
if contInput.keys in ['', [], None]:  # No response was made
   contInput.keys=None
# store data for thisExp (ExperimentHandler)
thisExp.addData('contInput.keys',contInput.keys)
if contInput.keys != None:  # we had a response
    thisExp.addData('contInput.rt', contInput.rt)
thisExp.nextEntry()

# the Routine "instructPage" was not non-slip safe, so reset the non-slip timer
routineTimer.reset()

# set up handler to look after randomisation of conditions etc
repeat = data.TrialHandler(nReps=6, method='sequential', 
    extraInfo=expInfo, originPath=None,
    trialList=[None],
    seed=None, name='repeat')
thisExp.addLoop(repeat)  # add the loop to the experiment
thisRepeat = repeat.trialList[0]  # so we can initialise stimuli with some values
# abbreviate parameter names if possible (e.g. rgb=thisRepeat.rgb)
if thisRepeat != None:
    for paramName in thisRepeat.keys():
        exec(paramName + '= thisRepeat.' + paramName)

for thisRepeat in repeat:
    currentLoop = repeat
    # abbreviate parameter names if possible (e.g. rgb = thisRepeat.rgb)
    if thisRepeat != None:
        for paramName in thisRepeat.keys():
            exec(paramName + '= thisRepeat.' + paramName)
    
    #------Prepare to start Routine "trial"-------
    t = 0
    trialClock.reset()  # clock 
    frameN = -1
    routineTimer.add(10.000000)
    # update component parameters for each repeat
    #---The following code the main implementation of ths Experiment
    
    #--This part extracts parameter values which are defined using the buider mode.
    assignOrder = trialOrder[repeat.thisN] #-Helps decide the random order to be used on this repeat.
    BComp = expInfo['Boolean_Complexity'] #-Gets the Boolean complexity given by the test taker at the begining of the experiment.
    
    #--This part intiializes variable value which have to consistent throughout the experiment.
    ovalSize = [0.2, 0.15]
    #shpSize = 
    stmdepth = -20.0
    assignTrack    = []
    
    
    #fshape      = ['square', 'triangle']
    #fnumber     = ['2','1']
    #fstcolor    = ['black', 'white']
    #finCat      = ['0','1']
    fparity     = 1
    #
    ##--This part creates the randomization of the property value assignment.
    #shuffle(fshape)
    #shuffle(fnumber)
    #shuffle(fstcolor)
    #shuffle(finCat)
    
    typeNo = testTypeNo[repeat.thisN]
    print ['typeNp', typeNo]
    
    if typeNo == 1:
        shpID   = [0, 1]
        numID   = [0, 1]
        colID   = [0, 1]
        inCatID = [0, 1]
    elif typeNo == 2:
        shpID   = [0, 1] 
        numID   = [1, 0] ##
        colID   = [0, 1]
        inCatID = [0, 1]
    elif typeNo == 3:
        shpID   = [0, 1]
        numID   = [0, 1]
        colID   = [1, 0] ##
        inCatID = [0, 1]
    elif typeNo == 4:
        shpID   = [0, 1]
        numID   = [0, 1]
        colID   = [0, 1] 
        inCatID = [1, 0] ##
    elif typeNo == 5:
        shpID   = [0, 1]
        numID   = [1, 0] #
        colID   = [1, 0] #
        inCatID = [0, 1]
    elif typeNo == 6:
        shpID   = [0, 1]
        numID   = [1, 0] #
        colID   = [0, 1]
        inCatID = [1, 0] #
    elif typeNo == 7:
        shpID   = [0, 1]
        numID   = [0, 1]
        colID   = [1, 0] #
        inCatID = [1, 0] #
    elif typeNo == 8:
        shpID   = [0, 1]
        numID   = [1, 0] #
        colID   = [1, 0] #
        inCatID = [1, 0] #
    elif typeNo == 9:
        shpID   = [1, 0] #
        numID   = [0, 1]
        colID   = [0, 1]
        inCatID = [0, 1]
    elif typeNo == 10:
        shpID   = [1, 0] #
        numID   = [1, 0] #
        colID   = [0, 1]
        inCatID = [0, 1]
    elif typeNo == 11:
        shpID   = [1, 0] #
        numID   = [0, 1]
        colID   = [1, 0] ##
        inCatID = [0, 1]
    elif typeNo == 12:
        shpID   = [1, 0]
        numID   = [0, 1]
        colID   = [0, 1] 
        inCatID = [1, 0] ##
    elif typeNo == 13:
        shpID   = [1, 0]
        numID   = [1, 0] #
        colID   = [1, 0] #
        inCatID = [0, 1]
    elif typeNo == 14:
        shpID   = [1, 0]
        numID   = [1, 0] #
        colID   = [0, 1]
        inCatID = [1, 0] #
    elif typeNo == 15:
        shpID   = [1, 0]
        numID   = [0, 1]
        colID   = [1, 0] #
        inCatID = [1, 0] #
    elif typeNo == 16:
        shpID   = [1, 0]
        numID   = [1, 0] #
        colID   = [1, 0] #
        inCatID = [1, 0] #
    
    
    fshape = []
    fshape.append( shapeList[ shpID[0] ] )
    fshape.append( shapeList[ shpID[1] ] )
    
    fnumber = []
    fnumber.append( numberList[ numID[0] ] )
    fnumber.append( numberList[ numID[1] ] )
    
    fstcolor = []
    fstcolor.append( stcolorList[ colID[0] ] )
    fstcolor.append( stcolorList[ colID[1] ] )
    
    finCat = []
    finCat.append( inCatList[ inCatID[0] ] )
    finCat.append( inCatList[ inCatID[1] ] )
    
    print fshape
    print fnumber
    print fstcolor
    print finCat
    
    
    #--The Shape, Color and Number Properties will be randomly assigned to a, b and c.
    #--a,b and c will be used instead of property lists above while creating the parity clusters.
    #--Use of a, b and c will allow us to use a standard formula type approach to create the clusters
    #-- while maintining the need for random use of properties to give the desired Boolean Complexity.
    a = [None, None]
    b = [None, None]
    c = [None, None]
    
    #--The Following part will randomly assign 
    #-- the property lists of Shape, Color and Number to lists a, b and c.
    #-- while allowing us to track the assignemnts.
    #--variable assignOrder is used for this purpose.
    #-- Note: This method is usable only because we have small number of possible random assignments. 
    #--       If the number of Randomizations have increased and tracking has become more complex I recommend using a new method to do this.
    
    #--This Part is used to perform the random assignment of the properties to a, b, and c.
    if assignOrder == 1:
        a = fshape
        b = fnumber
        c = fstcolor
        finCat = ['0','1']
        fparity = 1
        assignOrder = 1
    elif assignOrder == 2:
        a = fshape
        b = fnumber
        c = fstcolor
        finCat = ['1','0']
        fparity = 0
        assignOrder = 2
    elif assignOrder == 3:
        a = fshape
        b = fstcolor
        c = fnumber
        finCat = ['0','1']
        fparity = 1
        assignOrder = 3
    elif assignOrder == 4:
        a = fshape
        b = fstcolor
        c = fnumber
        finCat = ['1','0']
        fparity = 0
        assignOrder = 4
    elif assignOrder == 5:
        a = fnumber
        b = fshape
        c = fstcolor
        finCat = ['0','1']
        fparity = 1
        assignOrder = 5
    elif assignOrder == 6:
        a = fnumber
        b = fshape
        c = fstcolor
        finCat = ['1','0']
        fparity = 0
        assignOrder = 6
    elif assignOrder == 7:
        a = fnumber
        b = fstcolor
        c = fshape
        finCat = ['0','1']
        fparity = 1
        assignOrder = 7
    elif assignOrder == 8:
        a = fnumber
        b = fstcolor
        c = fshape
        finCat = ['1','0']
        fparity = 0
        assignOrder = 8
    elif assignOrder == 9:
        a = fstcolor
        b = fnumber
        c = fshape
        finCat = ['0','1']
        fparity = 1
        assignOrder = 9
    elif assignOrder == 10:
        a = fstcolor
        b = fnumber
        c = fshape
        finCat = ['1','0']
        fparity = 0
        assignOrder = 10
    elif assignOrder == 11:
        a = fstcolor
        b = fshape
        c = fnumber
        finCat = ['0','1']
        fparity = 1
        assignOrder = 11
    else:
        a = fstcolor
        b = fshape
        c = fnumber
        finCat = ['1','0']
        fparity = 0
        assignOrder = 12
    
    #--This part will create the stimuli objects with discrete property value combinations.
    stimuli = [] #-- we will use this list to perform the operations commons to all the stimuli objects.
    
    if BComp == '2': #---Assignment for Boolean Complexity 2
        stimuli.append(Stimulus(a[1], b[1], c[1], finCat[1], fparity, assignOrder))
        stimuli.append(Stimulus(a[1], b[1], c[0], finCat[1], fparity, assignOrder))
        stimuli.append(Stimulus(a[0], b[0], c[0], finCat[0], fparity, assignOrder))
        stimuli.append(Stimulus(a[0], b[0], c[1], finCat[0], fparity, assignOrder))
        stimuli.append(Stimulus(a[0], b[1], c[0], finCat[0], fparity, assignOrder))
        stimuli.append(Stimulus(a[0], b[1], c[1], finCat[0], fparity, assignOrder))
        stimuli.append(Stimulus(a[1], b[0], c[1], finCat[0], fparity, assignOrder))
        stimuli.append(Stimulus(a[1], b[0], c[0], finCat[0], fparity, assignOrder))
    
    if BComp == '5': #---Assignment for Boolean Complexity 5
        stimuli.append(Stimulus(a[1], b[1], c[1], finCat[1], fparity, assignOrder))
        stimuli.append(Stimulus(a[1], b[0], c[0], finCat[1], fparity, assignOrder))
        stimuli.append(Stimulus(a[0], b[0], c[0], finCat[0], fparity, assignOrder))
        stimuli.append(Stimulus(a[0], b[0], c[1], finCat[0], fparity, assignOrder))
        stimuli.append(Stimulus(a[0], b[1], c[0], finCat[0], fparity, assignOrder))
        stimuli.append(Stimulus(a[0], b[1], c[1], finCat[0], fparity, assignOrder))
        stimuli.append(Stimulus(a[1], b[0], c[1], finCat[0], fparity, assignOrder))
        stimuli.append(Stimulus(a[1], b[1], c[0], finCat[0], fparity, assignOrder))
    
    if BComp == '6': #---Assignment for Boolean Complexity 6
        stimuli.append(Stimulus(a[1], b[1], c[1], finCat[1], fparity, assignOrder))
        stimuli.append(Stimulus(a[0], b[0], c[0], finCat[1], fparity, assignOrder))
        stimuli.append(Stimulus(a[1], b[1], c[0], finCat[0], fparity, assignOrder))
        stimuli.append(Stimulus(a[0], b[0], c[1], finCat[0], fparity, assignOrder))
        stimuli.append(Stimulus(a[0], b[1], c[0], finCat[0], fparity, assignOrder))
        stimuli.append(Stimulus(a[0], b[1], c[1], finCat[0], fparity, assignOrder))
        stimuli.append(Stimulus(a[1], b[0], c[1], finCat[0], fparity, assignOrder))
        stimuli.append(Stimulus(a[1], b[0], c[0], finCat[0], fparity, assignOrder))
    
    #--Here we right the stimuli element property values to 'stimuliList.xlsx'
    #--This excel file will be used as input during testing.
    r = 1
    shuffle(stimuli)
    for stm in stimuli:
        wb = load_workbook(filename = 'categoryStimuliList.xlsx')
        ws = wb.get_sheet_by_name("Sheet1")
        
        ws.cell(row=r,column=0).value = stm.s #--shape value
        ws.cell(row=r,column=1).value = stm.n #--number value
        ws.cell(row=r,column=2).value = stm.c #--color value
        ws.cell(row=r,column=3).value = stm.cat #--category assignment value
        if stm.cat == '1':
            ws.cell(row=r,column=4).value = 'z' #--correct answer key value assignment
        else:
            ws.cell(row=r,column=4).value = 'm' #--correct answer key value assignment
        ws.cell(row=r,column=5).value = stm.p #--parity value assignment.
        wb.save('categoryStimuliList.xlsx')
        r = r+1 #--incrementing r value so that next set of vaues can be written to the next line in the excel file.
    
    #--Displaying the Stimuli as 'in category' or 'not in category'
    #--the following variables will be used in displaying the stimuli
    stimuliList = []
    stmNum = -1
    #--In Category stimuli position variables
    inX = -0.375
    inY = 0.375
    #--Not In Category stimuli position variables
    outX = -0.375
    outY = -0.125
    
    shuffle(stimuli)
    
    #---Creating Stimuli-1
    #--This Part decides the positioning of the stimuli based on its parity
    stmNum  = stmNum+1
    #--Assinging the center point for the stimuli. 
    if stimuli[stmNum].cat == '1':
        if inX > 0.5:
            inX = -0.375
            inY = 0.125
        posX = inX
        inX = inX+0.25
        posY = inY
    else:
        if outX > 0.5:
            outX = -0.375
            outY = -0.375
        posX = outX
        outX = outX+0.25
        posY = outY
    #--Drawing The Oval
    oval1 = visual.Polygon(win=win,
        edges = 100, size=ovalSize,
        ori=0, pos=[posX, posY],
        lineWidth=3, lineColor=[-1,-1,-1], lineColorSpace='rgb',
        fillColor=[1,1,1], fillColorSpace='rgb',
        opacity=1,depth=stmdepth+1.0, 
        interpolate=True)
    oval1.setAutoDraw(True, log=None)
    stimuliList.append(oval1)
    #--This part Creates the shapes 
    #--This part draws triangles
    if stimuli[stmNum].s == 'triangle':
        if stimuli[stmNum].n == '2': #--controls whether to show 2 elements or 1.
            shp11 = visual.Polygon(win, edges=3, radius=1, size=[0.015,0.035], lineColor=[-1,-1,-1],
                pos=[posX-0.04, posY], fillColor=stimuli[stmNum].c, lineWidth=2, autoLog=False,
                ori = 0, depth=stmdepth)
            shp11.tStart = t
            shp11.frameNStart = frameN
            shp11.setAutoDraw(True, log=None)
            stimuliList.append(shp11)
    
        shp12 = visual.Polygon(win, edges=3, radius=1, size=[0.015,0.035], lineColor=[-1,-1,-1],
            pos=[posX, posY], fillColor=stimuli[stmNum].c, lineWidth=2, autoLog=False,
            ori = 0, depth=stmdepth)
        shp12.tStart = t
        shp12.frameNStart = frameN
        shp12.setAutoDraw(True, log=None)
        stimuliList.append(shp12)
    
    else:
    #--This part draws squares
        if stimuli[stmNum].n == '2': #--controls whether to show 2 elements or 1.
            shp11 = visual.Rect(win=win,
                width=[0.025, 0.05][0], height=[0.025, 0.05][1],
                ori=0, pos=[posX-0.04, posY],
                lineWidth=2, lineColor=[-1,-1,-1], lineColorSpace='rgb',
                fillColor=stimuli[stmNum].c, fillColorSpace='rgb',
                opacity=1,depth=stmdepth, 
                interpolate=True)
            shp11.tStart = t
            shp11.frameNStart = frameN
            shp11.setAutoDraw(True, log=None)
            stimuliList.append(shp11)
    
        shp12 = visual.Rect(win=win,
            width=[0.025, 0.05][0], height=[0.025, 0.05][1],
            ori=0, pos=[posX, posY],
            lineWidth=2, lineColor=[-1,-1,-1], lineColorSpace='rgb',
            fillColor=stimuli[stmNum].c, fillColorSpace='rgb',
            opacity=1,depth=stmdepth, 
            interpolate=True)
        shp12.tStart = t
        shp12.frameNStart = frameN
        shp12.setAutoDraw(True, log=None)
        stimuliList.append(shp12)
    
    #---creating Stimuli-2
    #--This Part decides the positioning of the stimuli based on its parity
    stmNum  = stmNum+1
    #--Assinging the center point for the stimuli.
    if stimuli[stmNum].cat == '1':
        if inX > 0.5:
            inX = -0.375
            inY = 0.125
        posX = inX
        inX = inX+0.25
        posY = inY
    else:
        if outX > 0.5:
            outX = -0.375
            outY = -0.375
        posX = outX
        outX = outX+0.25
        posY = outY
    #--This draws the oval.
    oval2 = visual.Polygon(win=win,
        edges = 100, size=ovalSize,
        ori=0, pos=[posX, posY],
        lineWidth=3, lineColor=[-1,-1,-1], lineColorSpace='rgb',
        fillColor=[1,1,1], fillColorSpace='rgb',
        opacity=1,depth=stmdepth+1.0, 
        interpolate=True)
    oval2.setAutoDraw(True, log=None)
    stimuliList.append(oval2)
    #--This part creates the triangles
    if stimuli[stmNum].s == 'triangle':
        if stimuli[stmNum].n == '2': #--controls whether to show 2 elements or 1.
            shp21 = visual.Polygon(win, edges=3, radius=1, size=[0.015,0.035], lineColor=[-1,-1,-1],
                pos=[posX-0.04, posY], fillColor=stimuli[stmNum].c, lineWidth=2, autoLog=False,
                ori = 0, depth=stmdepth)
            shp21.tStart = t
            shp21.frameNStart = frameN
            shp21.setAutoDraw(True, log=None)
            stimuliList.append(shp21)
    
        shp22 = visual.Polygon(win, edges=3, radius=1, size=[0.015,0.035], lineColor=[-1,-1,-1],
            pos=[posX, posY], fillColor=stimuli[stmNum].c, lineWidth=2, autoLog=False,
            ori = 0, depth=stmdepth)
        shp22.tStart = t
        shp22.frameNStart = frameN
        shp22.setAutoDraw(True, log=None)
        stimuliList.append(shp22)
    
    else:
    #--This part creates the sqaures.
        if stimuli[stmNum].n == '2': #--controls whether to show 2 elements or 1.
            shp21 = visual.Rect(win=win,
                width=[0.025, 0.05][0], height=[0.025, 0.05][1],
                ori=0, pos=[posX-0.04, posY],
                lineWidth=2, lineColor=[-1,-1,-1], lineColorSpace='rgb',
                fillColor=stimuli[stmNum].c, fillColorSpace='rgb',
                opacity=1,depth=stmdepth, 
                interpolate=True)
            shp21.tStart = t
            shp21.frameNStart = frameN
            shp21.setAutoDraw(True, log=None)
            stimuliList.append(shp21)
    
        shp22 = visual.Rect(win=win,
            width=[0.025, 0.05][0], height=[0.025, 0.05][1],
            ori=0, pos=[posX, posY],
            lineWidth=2, lineColor=[-1,-1,-1], lineColorSpace='rgb',
            fillColor=stimuli[stmNum].c, fillColorSpace='rgb',
            opacity=1,depth=stmdepth, 
            interpolate=True)
        shp22.tStart = t
        shp22.frameNStart = frameN
        shp22.setAutoDraw(True, log=None)
        stimuliList.append(shp22)
    
    #   creating Stimuli-3
    #This Part decides the positioning of the stimuli based on its parity
    stmNum  = stmNum+1
    #--Assinging the center point for the stimuli.
    if stimuli[stmNum].cat == '1':
        if inX > 0.5:
            inX = -0.375
            inY = 0.125
        posX = inX
        inX = inX+0.25
        posY = inY
    else:
        if outX > 0.5:
            outX = -0.375
            outY = -0.375
        posX = outX
        outX = outX+0.25
        posY = outY
    #--This draws the ovals
    oval3 = visual.Polygon(win=win,
        edges = 100, size=ovalSize,
        ori=0, pos=[posX, posY],
        lineWidth=3, lineColor=[-1,-1,-1], lineColorSpace='rgb',
        fillColor=[1,1,1], fillColorSpace='rgb',
        opacity=1,depth=stmdepth+1.0, 
        interpolate=True)
    oval3.setAutoDraw(True, log=None)
    stimuliList.append(oval3)
    #--This part draws the traiangle
    if stimuli[stmNum].s == 'triangle':
        if stimuli[stmNum].n == '2': #--controls whether to show 2 elements or 1.
            shp31 = visual.Polygon(win, edges=3, radius=1, size=[0.015,0.035], lineColor=[-1,-1,-1],
                pos=[posX-0.04, posY], fillColor=stimuli[stmNum].c, lineWidth=2, autoLog=False,
                ori = 0, depth=stmdepth)
            shp31.tStart = t
            shp31.frameNStart = frameN
            shp31.setAutoDraw(True, log=None)
            stimuliList.append(shp31)
    
        shp32 = visual.Polygon(win, edges=3, radius=1, size=[0.015,0.035], lineColor=[-1,-1,-1],
            pos=[posX, posY], fillColor=stimuli[stmNum].c, lineWidth=2, autoLog=False,
            ori = 0, depth=stmdepth)
        shp32.tStart = t
        shp32.frameNStart = frameN
        shp32.setAutoDraw(True, log=None)
        stimuliList.append(shp32)
    
    else:
    #-- This part draws squares
        if stimuli[stmNum].n == '2': #--controls whether to show 2 elements or 1.
            shp31 = visual.Rect(win=win,
                width=[0.025, 0.05][0], height=[0.025, 0.05][1],
                ori=0, pos=[posX-0.04, posY],
                lineWidth=2, lineColor=[-1,-1,-1], lineColorSpace='rgb',
                fillColor=stimuli[stmNum].c, fillColorSpace='rgb',
                opacity=1,depth=stmdepth, 
                interpolate=True)
            shp31.tStart = t
            shp31.frameNStart = frameN
            shp31.setAutoDraw(True, log=None)
            stimuliList.append(shp31)
    
        shp32 = visual.Rect(win=win,
            width=[0.025, 0.05][0], height=[0.025, 0.05][1],
            ori=0, pos=[posX, posY],
            lineWidth=2, lineColor=[-1,-1,-1], lineColorSpace='rgb',
            fillColor=stimuli[stmNum].c, fillColorSpace='rgb',
            opacity=1,depth=stmdepth, 
            interpolate=True)
        shp32.tStart = t
        shp32.frameNStart = frameN
        shp32.setAutoDraw(True, log=None)
        stimuliList.append(shp32)
    
    #   creating Stimuli-4
    #This Part decides the positioning of the stimuli based on its parity
    stmNum  = stmNum+1
    #--Assinging the center point for the stimuli.
    if stimuli[stmNum].cat == '1':
        if inX > 0.5:
            inX = -0.375
            inY = 0.125
        posX = inX
        inX = inX+0.25
        posY = inY
    else:
        if outX > 0.5:
            outX = -0.375
            outY = -0.375
        posX = outX
        outX = outX+0.25
        posY = outY
    #--This part draws the oval
    oval4 = visual.Polygon(win=win,
        edges = 100, size=ovalSize,
        ori=0, pos=[posX, posY],
        lineWidth=3, lineColor=[-1,-1,-1], lineColorSpace='rgb',
        fillColor=[1,1,1], fillColorSpace='rgb',
        opacity=1,depth=stmdepth+1.0, 
        interpolate=True)
    oval4.setAutoDraw(True, log=None)
    stimuliList.append(oval4)
    #--This part draws the triangle
    if stimuli[stmNum].s == 'triangle':
        if stimuli[stmNum].n == '2': #--controls whether to show 2 elements or 1.
            shp41 = visual.Polygon(win, edges=3, radius=1, size=[0.015,0.035], lineColor=[-1,-1,-1],
                pos=[posX-0.04, posY], fillColor=stimuli[stmNum].c, lineWidth=2, autoLog=False,
                ori = 0, depth=stmdepth)
            shp41.tStart = t
            shp41.frameNStart = frameN
            shp41.setAutoDraw(True, log=None)
            stimuliList.append(shp41)
    
        shp42 = visual.Polygon(win, edges=3, radius=1, size=[0.015,0.035], lineColor=[-1,-1,-1],
            pos=[posX, posY], fillColor=stimuli[stmNum].c, lineWidth=2, autoLog=False,
            ori = 0, depth=stmdepth)
        shp42.tStart = t
        shp42.frameNStart = frameN
        shp42.setAutoDraw(True, log=None)
        stimuliList.append(shp42)
    
    else:
    #This Part draws the sqaures.
        if stimuli[stmNum].n == '2': #--controls whether to show 2 elements or 1.
            shp41 = visual.Rect(win=win,
                width=[0.025, 0.05][0], height=[0.025, 0.05][1],
                ori=0, pos=[posX-0.04, posY],
                lineWidth=2, lineColor=[-1,-1,-1], lineColorSpace='rgb',
                fillColor=stimuli[stmNum].c, fillColorSpace='rgb',
                opacity=1,depth=stmdepth, 
                interpolate=True)
            shp41.tStart = t
            shp41.frameNStart = frameN
            shp41.setAutoDraw(True, log=None)
            stimuliList.append(shp41)
    
        shp42 = visual.Rect(win=win,
            width=[0.025, 0.05][0], height=[0.025, 0.05][1],
            ori=0, pos=[posX, posY],
            lineWidth=2, lineColor=[-1,-1,-1], lineColorSpace='rgb',
            fillColor=stimuli[stmNum].c, fillColorSpace='rgb',
            opacity=1,depth=stmdepth, 
            interpolate=True)
        shp42.tStart = t
        shp42.frameNStart = frameN
        shp42.setAutoDraw(True, log=None)
        stimuliList.append(shp42)
    
    #--creating Stimuli-5
    #--This Part decides the positioning of the stimuli based on its parity
    stmNum  = stmNum+1
    #--Assinging the center point for the stimuli.
    if stimuli[stmNum].cat == '1':
        if inX > 0.5:
            inX = -0.375
            inY = 0.125
        posX = inX
        inX = inX+0.25
        posY = inY
    else:
        if outX > 0.5:
            outX = -0.375
            outY = -0.375
        posX = outX
        outX = outX+0.25
        posY = outY
    #--This part draws the oval
    oval5 = visual.Polygon(win=win,
        edges = 100, size=ovalSize,
        ori=0, pos=[posX, posY],
        lineWidth=3, lineColor=[-1,-1,-1], lineColorSpace='rgb',
        fillColor=[1,1,1], fillColorSpace='rgb',
        opacity=1,depth=stmdepth+1.0, 
        interpolate=True)
    oval5.setAutoDraw(True, log=None)
    stimuliList.append(oval5)
    #--This part draws the Triangle
    if stimuli[stmNum].s == 'triangle':
        if stimuli[stmNum].n == '2': #--controls whether to show 2 elements or 1.
            shp51 = visual.Polygon(win, edges=3, radius=1, size=[0.015,0.035], lineColor=[-1,-1,-1],
                pos=[posX-0.04, posY], fillColor=stimuli[stmNum].c, lineWidth=2, autoLog=False,
                ori = 0, depth=stmdepth)
            shp51.tStart = t
            shp51.frameNStart = frameN
            shp51.setAutoDraw(True, log=None)
            stimuliList.append(shp51)
    
        shp52 = visual.Polygon(win, edges=3, radius=1, size=[0.015,0.035], lineColor=[-1,-1,-1],
            pos=[posX, posY], fillColor=stimuli[stmNum].c, lineWidth=2, autoLog=False,
            ori = 0, depth=stmdepth)
        shp52.tStart = t
        shp52.frameNStart = frameN
        shp52.setAutoDraw(True, log=None)
        stimuliList.append(shp52)
    
    else:
    #--This part draws the sqaures
        if stimuli[stmNum].n == '2': #--controls whether to show 2 elements or 1.
            shp51 = visual.Rect(win=win,
                width=[0.025, 0.05][0], height=[0.025, 0.05][1],
                ori=0, pos=[posX-0.04, posY],
                lineWidth=2, lineColor=[-1,-1,-1], lineColorSpace='rgb',
                fillColor=stimuli[stmNum].c, fillColorSpace='rgb',
                opacity=1,depth=stmdepth, 
                interpolate=True)
            shp51.tStart = t
            shp51.frameNStart = frameN
            shp51.setAutoDraw(True, log=None)
            stimuliList.append(shp51)
    
        shp52 = visual.Rect(win=win,
            width=[0.025, 0.05][0], height=[0.025, 0.05][1],
            ori=0, pos=[posX, posY],
            lineWidth=2, lineColor=[-1,-1,-1], lineColorSpace='rgb',
            fillColor=stimuli[stmNum].c, fillColorSpace='rgb',
            opacity=1,depth=stmdepth, 
            interpolate=True)
        shp52.tStart = t
        shp52.frameNStart = frameN
        shp52.setAutoDraw(True, log=None)
        stimuliList.append(shp52)
    
    #   creating Stimuli-6
    #This Part decides the positioning of the stimuli based on its parity
    stmNum  = stmNum+1
    #--Assinging the center point for the stimuli.
    if stimuli[stmNum].cat == '1':
        if inX > 0.5:
            inX = -0.375
            inY = 0.125
        posX = inX
        inX = inX+0.25
        posY = inY
    else:
        if outX > 0.5:
            outX = -0.375
            outY = -0.375
        posX = outX
        outX = outX+0.25
        posY = outY
    #--This part draws the Oval
    oval6 = visual.Polygon(win=win,
        edges = 100, size=ovalSize,
        ori=0, pos=[posX, posY],
        lineWidth=3, lineColor=[-1,-1,-1], lineColorSpace='rgb',
        fillColor=[1,1,1], fillColorSpace='rgb',
        opacity=1,depth=stmdepth+1.0, 
        interpolate=True)
    oval6.setAutoDraw(True, log=None)
    stimuliList.append(oval6)
    #--This part draws the Triangles
    if stimuli[stmNum].s == 'triangle':
        if stimuli[stmNum].n == '2': #--controls whether to show 2 elements or 1.
            shp61 = visual.Polygon(win, edges=3, radius=1, size=[0.015,0.035], lineColor=[-1,-1,-1],
                pos=[posX-0.04, posY], fillColor=stimuli[stmNum].c, lineWidth=2, autoLog=False,
                ori = 0, depth=stmdepth)
            shp61.tStart = t
            shp61.frameNStart = frameN
            shp61.setAutoDraw(True, log=None)
            stimuliList.append(shp61)
    
        shp62 = visual.Polygon(win, edges=3, radius=1, size=[0.015,0.035], lineColor=[-1,-1,-1],
            pos=[posX, posY], fillColor=stimuli[stmNum].c, lineWidth=2, autoLog=False,
            ori = 0, depth=stmdepth)
        shp62.tStart = t
        shp62.frameNStart = frameN
        shp62.setAutoDraw(True, log=None)
        stimuliList.append(shp62)
    
    else:
    #--This part draws the squares
        if stimuli[stmNum].n == '2': #--controls whether to show 2 elements or 1.
            shp61 = visual.Rect(win=win,
                width=[0.025, 0.05][0], height=[0.025, 0.05][1],
                ori=0, pos=[posX-0.04, posY],
                lineWidth=2, lineColor=[-1,-1,-1], lineColorSpace='rgb',
                fillColor=stimuli[stmNum].c, fillColorSpace='rgb',
                opacity=1,depth=stmdepth, 
                interpolate=True)
            shp61.tStart = t
            shp61.frameNStart = frameN
            shp61.setAutoDraw(True, log=None)
            stimuliList.append(shp61)
    
        shp62 = visual.Rect(win=win,
            width=[0.025, 0.05][0], height=[0.025, 0.05][1],
            ori=0, pos=[posX, posY],
            lineWidth=2, lineColor=[-1,-1,-1], lineColorSpace='rgb',
            fillColor=stimuli[stmNum].c, fillColorSpace='rgb',
            opacity=1,depth=stmdepth, 
            interpolate=True)
        shp62.tStart = t
        shp62.frameNStart = frameN
        shp62.setAutoDraw(True, log=None)
        stimuliList.append(shp62)
    
    #   creating Stimuli-7
    #This Part decides the positioning of the stimuli based on its parity
    stmNum  = stmNum+1
    #--Assinging the center point for the stimuli.
    if stimuli[stmNum].cat == '1':
        if inX > 0.5:
            inX = -0.375
            inY = 0.125
        posX = inX
        inX = inX+0.25
        posY = inY
    else:
        if outX > 0.5:
            outX = -0.375
            outY = -0.375
        posX = outX
        outX = outX+0.25
        posY = outY
    #--This part draws the Ovals
    oval7 = visual.Polygon(win=win,
        edges = 100, size=ovalSize,
        ori=0, pos=[posX, posY],
        lineWidth=3, lineColor=[-1,-1,-1], lineColorSpace='rgb',
        fillColor=[1,1,1], fillColorSpace='rgb',
        opacity=1,depth=stmdepth+1.0, 
        interpolate=True)
    oval7.setAutoDraw(True, log=None)
    stimuliList.append(oval7)
    #--This part draws the Triangles
    if stimuli[stmNum].s == 'triangle':
        if stimuli[stmNum].n == '2': #--controls whether to show 2 elements or 1.
            shp71 = visual.Polygon(win, edges=3, radius=1, size=[0.015,0.035], lineColor=[-1,-1,-1],
                pos=[posX-0.04, posY], fillColor=stimuli[stmNum].c, lineWidth=2, autoLog=False,
                ori = 0, depth=stmdepth)
            shp71.tStart = t
            shp71.frameNStart = frameN
            shp71.setAutoDraw(True, log=None)
            stimuliList.append(shp71)
    
        shp72 = visual.Polygon(win, edges=3, radius=1, size=[0.015,0.035], lineColor=[-1,-1,-1],
            pos=[posX, posY], fillColor=stimuli[stmNum].c, lineWidth=2, autoLog=False,
            ori = 0, depth=stmdepth)
        shp72.tStart = t
        shp72.frameNStart = frameN
        shp72.setAutoDraw(True, log=None)
        stimuliList.append(shp72)
    
    else:
    #--This part draws the sqaures
        if stimuli[stmNum].n == '2': #--controls whether to show 2 elements or 1.
            shp71 = visual.Rect(win=win,
                width=[0.025, 0.05][0], height=[0.025, 0.05][1],
                ori=0, pos=[posX-0.04, posY],
                lineWidth=2, lineColor=[-1,-1,-1], lineColorSpace='rgb',
                fillColor=stimuli[stmNum].c, fillColorSpace='rgb',
                opacity=1,depth=stmdepth, 
                interpolate=True)
            shp71.tStart = t
            shp71.frameNStart = frameN
            shp71.setAutoDraw(True, log=None)
            stimuliList.append(shp71)
    
        shp72 = visual.Rect(win=win,
            width=[0.025, 0.05][0], height=[0.025, 0.05][1],
            ori=0, pos=[posX, posY],
            lineWidth=2, lineColor=[-1,-1,-1], lineColorSpace='rgb',
            fillColor=stimuli[stmNum].c, fillColorSpace='rgb',
            opacity=1,depth=stmdepth, 
            interpolate=True)
        shp72.tStart = t
        shp72.frameNStart = frameN
        shp72.setAutoDraw(True, log=None)
        stimuliList.append(shp72)
    
    #   creating Stimuli-8
    #This Part decides the positioning of the stimuli based on its parity
    stmNum  = stmNum+1
    #--Assinging the center point for the stimuli.
    if stimuli[stmNum].cat == '1':
        if inX >0.5:
            inX = -0.375
            inY = 0.125
        posX = inX
        inX = inX+0.25
        posY = inY
    else:
        if outX > 0.5:
            outX = -0.375
            outY = -0.375
        posX = outX
        outX = outX+0.25
        posY = outY
    #This part draws the Oval
    oval8 = visual.Polygon(win=win,
        edges = 100, size=ovalSize,
        ori=0, pos=[posX, posY],
        lineWidth=3, lineColor=[-1,-1,-1], lineColorSpace='rgb',
        fillColor=[1,1,1], fillColorSpace='rgb',
        opacity=1,depth=stmdepth, 
        interpolate=True)
    oval8.setAutoDraw(True, log=None)
    stimuliList.append(oval8)
    #--This part draws the Triangles
    if stimuli[stmNum].s == 'triangle':
        if stimuli[stmNum].n == '2': #--controls whether to show 2 elements or 1.
            shp81 = visual.Polygon(win, edges=3, radius=1, size=[0.015,0.035], lineColor=[-1,-1,-1],
                pos=[posX-0.04, posY], fillColor=stimuli[stmNum].c, lineWidth=2, autoLog=False,
                ori = 0, depth=stmdepth)
            shp81.tStart = t
            shp81.frameNStart = frameN
            shp81.setAutoDraw(True, log=None)
            stimuliList.append(shp81)
    
        shp82 = visual.Polygon(win, edges=3, radius=1, size=[0.015,0.035], lineColor=[-1,-1,-1],
            pos=[posX, posY], fillColor=stimuli[stmNum].c, lineWidth=2, autoLog=False,
            ori = 0, depth=stmdepth)
        shp82.tStart = t
        shp82.frameNStart = frameN
        shp82.setAutoDraw(True, log=None)
        stimuliList.append(shp82)
    
    else:
    #--This part draws the squares.
        if stimuli[stmNum].n == '2': #--controls whether to show 2 elements or 1.
            shp81 = visual.Rect(win=win,
                width=[0.025, 0.05][0], height=[0.025, 0.05][1],
                ori=0, pos=[posX-0.04, posY],
                lineWidth=2, lineColor=[-1,-1,-1], lineColorSpace='rgb',
                fillColor=stimuli[stmNum].c, fillColorSpace='rgb',
                opacity=1,depth=stmdepth,
                interpolate=True)
            shp81.tStart = t
            shp81.frameNStart = frameN
            shp81.setAutoDraw(True, log=None)
            stimuliList.append(shp81)
    
        shp82 = visual.Rect(win=win,
            width=[0.025, 0.05][0], height=[0.025, 0.05][1],
            ori=0, pos=[posX, posY],
            lineWidth=2, lineColor=[-1,-1,-1], lineColorSpace='rgb',
            fillColor=stimuli[stmNum].c, fillColorSpace='rgb',
            opacity=1,depth=stmdepth,
            interpolate=True)
        shp82.tStart = t
        shp82.frameNStart = frameN
        shp82.setAutoDraw(True, log=None)
        stimuliList.append(shp82)
    trialNum.setText(repeat.thisN+1
)
    # keep track of which components have finished
    trialComponents = []
    trialComponents.append(topTitle)
    trialComponents.append(inCatText)
    trialComponents.append(bottomTitle)
    trialComponents.append(notinCatText)
    trialComponents.append(trialmsg)
    trialComponents.append(trialNum)
    trialComponents.append(ISI)
    trialComponents.append(categories)
    trialComponents.append(categorySeparator)
    for thisComponent in trialComponents:
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    
    #-------Start Routine "trial"-------
    continueRoutine = True
    while continueRoutine and routineTimer.getTime() > 0:
        # get current time
        t = trialClock.getTime()
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        
        # *topTitle* updates
        if t >= 0.0 and topTitle.status == NOT_STARTED:
            # keep track of start time/frame for later
            topTitle.tStart = t  # underestimates by a little under one frame
            topTitle.frameNStart = frameN  # exact frame index
            topTitle.setAutoDraw(True)
        if topTitle.status == STARTED and t >= (0.0 + (10-win.monitorFramePeriod*0.75)): #most of one frame period left
            topTitle.setAutoDraw(False)
        
        # *inCatText* updates
        if t >= 0.0 and inCatText.status == NOT_STARTED:
            # keep track of start time/frame for later
            inCatText.tStart = t  # underestimates by a little under one frame
            inCatText.frameNStart = frameN  # exact frame index
            inCatText.setAutoDraw(True)
        if inCatText.status == STARTED and t >= (0.0 + (10-win.monitorFramePeriod*0.75)): #most of one frame period left
            inCatText.setAutoDraw(False)
        
        # *bottomTitle* updates
        if t >= 0.0 and bottomTitle.status == NOT_STARTED:
            # keep track of start time/frame for later
            bottomTitle.tStart = t  # underestimates by a little under one frame
            bottomTitle.frameNStart = frameN  # exact frame index
            bottomTitle.setAutoDraw(True)
        if bottomTitle.status == STARTED and t >= (0.0 + (10-win.monitorFramePeriod*0.75)): #most of one frame period left
            bottomTitle.setAutoDraw(False)
        
        # *notinCatText* updates
        if t >= 0.0 and notinCatText.status == NOT_STARTED:
            # keep track of start time/frame for later
            notinCatText.tStart = t  # underestimates by a little under one frame
            notinCatText.frameNStart = frameN  # exact frame index
            notinCatText.setAutoDraw(True)
        if notinCatText.status == STARTED and t >= (0.0 + (10-win.monitorFramePeriod*0.75)): #most of one frame period left
            notinCatText.setAutoDraw(False)
        
        # *trialmsg* updates
        if t >= 0.0 and trialmsg.status == NOT_STARTED:
            # keep track of start time/frame for later
            trialmsg.tStart = t  # underestimates by a little under one frame
            trialmsg.frameNStart = frameN  # exact frame index
            trialmsg.setAutoDraw(True)
        if trialmsg.status == STARTED and t >= (10-win.monitorFramePeriod*0.75): #most of one frame period left
            trialmsg.setAutoDraw(False)
        
        # *trialNum* updates
        if t >= 0.0 and trialNum.status == NOT_STARTED:
            # keep track of start time/frame for later
            trialNum.tStart = t  # underestimates by a little under one frame
            trialNum.frameNStart = frameN  # exact frame index
            trialNum.setAutoDraw(True)
        if trialNum.status == STARTED and t >= (10-win.monitorFramePeriod*0.75): #most of one frame period left
            trialNum.setAutoDraw(False)
        
        # *categories* updates
        if t >= 0.0 and categories.status == NOT_STARTED:
            # keep track of start time/frame for later
            categories.tStart = t  # underestimates by a little under one frame
            categories.frameNStart = frameN  # exact frame index
            categories.setAutoDraw(True)
        if categories.status == STARTED and t >= (10-win.monitorFramePeriod*0.75): #most of one frame period left
            categories.setAutoDraw(False)
        
        # *categorySeparator* updates
        if t >= 0.0 and categorySeparator.status == NOT_STARTED:
            # keep track of start time/frame for later
            categorySeparator.tStart = t  # underestimates by a little under one frame
            categorySeparator.frameNStart = frameN  # exact frame index
            categorySeparator.setAutoDraw(True)
        if categorySeparator.status == STARTED and t >= (10-win.monitorFramePeriod*0.75): #most of one frame period left
            categorySeparator.setAutoDraw(False)
        # *ISI* period
        if t >= 0.0 and ISI.status == NOT_STARTED:
            # keep track of start time/frame for later
            ISI.tStart = t  # underestimates by a little under one frame
            ISI.frameNStart = frameN  # exact frame index
            ISI.start(0.1-t)
        elif ISI.status == STARTED: #one frame should pass before updating params and completing
            ISI.complete() #finish the static period
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in trialComponents:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # check for quit (the Esc key)
        if endExpNow or event.getKeys(keyList=["escape"]):
            core.quit()
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    #-------Ending Routine "trial"-------
    for thisComponent in trialComponents:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    #--This part will delete all the displayed elements at the end of the Routine
    for stmComp in stimuliList:
        if hasattr(stmComp, "setAutoDraw"):
            stmComp.setAutoDraw(False)
    
    # set up handler to look after randomisation of conditions etc
    trials = data.TrialHandler(nReps=1, method='random', 
        extraInfo=expInfo, originPath=None,
        trialList=data.importConditions('categoryStimuliList.xlsx'),
        seed=None, name='trials')
    thisExp.addLoop(trials)  # add the loop to the experiment
    thisTrial = trials.trialList[0]  # so we can initialise stimuli with some values
    # abbreviate parameter names if possible (e.g. rgb=thisTrial.rgb)
    if thisTrial != None:
        for paramName in thisTrial.keys():
            exec(paramName + '= thisTrial.' + paramName)
    
    for thisTrial in trials:
        currentLoop = trials
        # abbreviate parameter names if possible (e.g. rgb = thisTrial.rgb)
        if thisTrial != None:
            for paramName in thisTrial.keys():
                exec(paramName + '= thisTrial.' + paramName)
        
        #------Prepare to start Routine "trialtest"-------
        t = 0
        trialtestClock.reset()  # clock 
        frameN = -1
        # update component parameters for each repeat
        #--The following variables control the display of elements during 
        #--the testing phase of the experiment.
        
        #--This controls the shape to be dispalyed
        nvertices = 0
        if shape == 'square':
            nvertices = 4
        else:
            nvertices = 3
        
        temp = nvertices
        
        #--This controls the number of shapes in the element.
        if number == 2:
            flag = 2
        if number == 3:
            flag = 3
        
        trialTestNum.setText(repeat.thisN+1)
        square1.setOpacity(number-1)
        square1.setFillColor(stcolor)
        square2.setFillColor(stcolor)
        triangle1.setOpacity(number-1)
        triangle1.setFillColor(stcolor)
        triangle2.setFillColor(stcolor)
        key_resp_3 = event.BuilderKeyResponse()  # create an object of type KeyResponse
        key_resp_3.status = NOT_STARTED
        # keep track of which components have finished
        trialtestComponents = []
        trialtestComponents.append(trialTestMsg)
        trialtestComponents.append(trialTestNum)
        trialtestComponents.append(oval)
        trialtestComponents.append(square1)
        trialtestComponents.append(square2)
        trialtestComponents.append(triangle1)
        trialtestComponents.append(triangle2)
        trialtestComponents.append(inputmsg)
        trialtestComponents.append(key_resp_3)
        for thisComponent in trialtestComponents:
            if hasattr(thisComponent, 'status'):
                thisComponent.status = NOT_STARTED
        
        #-------Start Routine "trialtest"-------
        continueRoutine = True
        while continueRoutine:
            # get current time
            t = trialtestClock.getTime()
            frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
            # update/draw components on each frame
            
            
            # *trialTestMsg* updates
            if t >= 0.0 and trialTestMsg.status == NOT_STARTED:
                # keep track of start time/frame for later
                trialTestMsg.tStart = t  # underestimates by a little under one frame
                trialTestMsg.frameNStart = frameN  # exact frame index
                trialTestMsg.setAutoDraw(True)
            
            # *trialTestNum* updates
            if t >= 0.0 and trialTestNum.status == NOT_STARTED:
                # keep track of start time/frame for later
                trialTestNum.tStart = t  # underestimates by a little under one frame
                trialTestNum.frameNStart = frameN  # exact frame index
                trialTestNum.setAutoDraw(True)
            
            # *oval* updates
            if t >= 0.0 and oval.status == NOT_STARTED:
                # keep track of start time/frame for later
                oval.tStart = t  # underestimates by a little under one frame
                oval.frameNStart = frameN  # exact frame index
                oval.setAutoDraw(True)
            
            # *square1* updates
            if (nvertices==4) and square1.status == NOT_STARTED:
                # keep track of start time/frame for later
                square1.tStart = t  # underestimates by a little under one frame
                square1.frameNStart = frameN  # exact frame index
                square1.setAutoDraw(True)
            
            # *square2* updates
            if (nvertices==4) and square2.status == NOT_STARTED:
                # keep track of start time/frame for later
                square2.tStart = t  # underestimates by a little under one frame
                square2.frameNStart = frameN  # exact frame index
                square2.setAutoDraw(True)
            
            # *triangle1* updates
            if (nvertices==3) and triangle1.status == NOT_STARTED:
                # keep track of start time/frame for later
                triangle1.tStart = t  # underestimates by a little under one frame
                triangle1.frameNStart = frameN  # exact frame index
                triangle1.setAutoDraw(True)
            
            # *triangle2* updates
            if (nvertices==3) and triangle2.status == NOT_STARTED:
                # keep track of start time/frame for later
                triangle2.tStart = t  # underestimates by a little under one frame
                triangle2.frameNStart = frameN  # exact frame index
                triangle2.setAutoDraw(True)
            
            # *inputmsg* updates
            if t >= 0.0 and inputmsg.status == NOT_STARTED:
                # keep track of start time/frame for later
                inputmsg.tStart = t  # underestimates by a little under one frame
                inputmsg.frameNStart = frameN  # exact frame index
                inputmsg.setAutoDraw(True)
            
            # *key_resp_3* updates
            if t >= 0.0 and key_resp_3.status == NOT_STARTED:
                # keep track of start time/frame for later
                key_resp_3.tStart = t  # underestimates by a little under one frame
                key_resp_3.frameNStart = frameN  # exact frame index
                key_resp_3.status = STARTED
                # keyboard checking is just starting
                key_resp_3.clock.reset()  # now t=0
                event.clearEvents(eventType='keyboard')
            if key_resp_3.status == STARTED:
                theseKeys = event.getKeys(keyList=['z', 'm'])
                
                # check for quit:
                if "escape" in theseKeys:
                    endExpNow = True
                if len(theseKeys) > 0:  # at least one key was pressed
                    key_resp_3.keys = theseKeys[-1]  # just the last key pressed
                    key_resp_3.rt = key_resp_3.clock.getTime()
                    # was this 'correct'?
                    if (key_resp_3.keys == str(corrAns)) or (key_resp_3.keys == corrAns):
                        key_resp_3.corr = 1
                    else:
                        key_resp_3.corr = 0
                    # a response ends the routine
                    continueRoutine = False
            
            # check if all components have finished
            if not continueRoutine:  # a component has requested a forced-end of Routine
                break
            continueRoutine = False  # will revert to True if at least one component still running
            for thisComponent in trialtestComponents:
                if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                    continueRoutine = True
                    break  # at least one component has not yet finished
            
            # check for quit (the Esc key)
            if endExpNow or event.getKeys(keyList=["escape"]):
                core.quit()
            
            # refresh the screen
            if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                win.flip()
        
        #-------Ending Routine "trialtest"-------
        for thisComponent in trialtestComponents:
            if hasattr(thisComponent, "setAutoDraw"):
                thisComponent.setAutoDraw(False)
        
        # check responses
        if key_resp_3.keys in ['', [], None]:  # No response was made
           key_resp_3.keys=None
           # was no response the correct answer?!
           if str(corrAns).lower() == 'none': key_resp_3.corr = 1  # correct non-response
           else: key_resp_3.corr = 0  # failed to respond (incorrectly)
        # store data for trials (TrialHandler)
        trials.addData('key_resp_3.keys',key_resp_3.keys)
        trials.addData('key_resp_3.corr', key_resp_3.corr)
        if key_resp_3.keys != None:  # we had a response
            trials.addData('key_resp_3.rt', key_resp_3.rt)
        # the Routine "trialtest" was not non-slip safe, so reset the non-slip timer
        routineTimer.reset()
        
        #------Prepare to start Routine "blankSpace"-------
        t = 0
        blankSpaceClock.reset()  # clock 
        frameN = -1
        routineTimer.add(0.250000)
        # update component parameters for each repeat
        # keep track of which components have finished
        blankSpaceComponents = []
        blankSpaceComponents.append(ISI_2)
        for thisComponent in blankSpaceComponents:
            if hasattr(thisComponent, 'status'):
                thisComponent.status = NOT_STARTED
        
        #-------Start Routine "blankSpace"-------
        continueRoutine = True
        while continueRoutine and routineTimer.getTime() > 0:
            # get current time
            t = blankSpaceClock.getTime()
            frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
            # update/draw components on each frame
            # *ISI_2* period
            if t >= 0.0 and ISI_2.status == NOT_STARTED:
                # keep track of start time/frame for later
                ISI_2.tStart = t  # underestimates by a little under one frame
                ISI_2.frameNStart = frameN  # exact frame index
                ISI_2.start(0.250)
            elif ISI_2.status == STARTED: #one frame should pass before updating params and completing
                ISI_2.complete() #finish the static period
            
            # check if all components have finished
            if not continueRoutine:  # a component has requested a forced-end of Routine
                break
            continueRoutine = False  # will revert to True if at least one component still running
            for thisComponent in blankSpaceComponents:
                if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                    continueRoutine = True
                    break  # at least one component has not yet finished
            
            # check for quit (the Esc key)
            if endExpNow or event.getKeys(keyList=["escape"]):
                core.quit()
            
            # refresh the screen
            if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                win.flip()
        
        #-------Ending Routine "blankSpace"-------
        for thisComponent in blankSpaceComponents:
            if hasattr(thisComponent, "setAutoDraw"):
                thisComponent.setAutoDraw(False)
        thisExp.nextEntry()
        
    # completed 1 repeats of 'trials'
    
    # get names of stimulus parameters
    if trials.trialList in ([], [None], None):  params = []
    else:  params = trials.trialList[0].keys()
    # save data for this loop
    trials.saveAsExcel(filename + '.xlsx', sheetName='trials',
        stimOut=params,
        dataOut=['n','all_mean','all_std', 'all_raw'])
    
    #------Prepare to start Routine "nextTrial"-------
    t = 0
    nextTrialClock.reset()  # clock 
    frameN = -1
    # update component parameters for each repeat
    contKeyPress = event.BuilderKeyResponse()  # create an object of type KeyResponse
    contKeyPress.status = NOT_STARTED
    # keep track of which components have finished
    nextTrialComponents = []
    nextTrialComponents.append(contMsg)
    nextTrialComponents.append(exitMsg)
    nextTrialComponents.append(contKeyPress)
    for thisComponent in nextTrialComponents:
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    
    #-------Start Routine "nextTrial"-------
    continueRoutine = True
    while continueRoutine:
        # get current time
        t = nextTrialClock.getTime()
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *contMsg* updates
        if t >= repeat.thisN < 5 and contMsg.status == NOT_STARTED:
            # keep track of start time/frame for later
            contMsg.tStart = t  # underestimates by a little under one frame
            contMsg.frameNStart = frameN  # exact frame index
            contMsg.setAutoDraw(True)
        
        # *exitMsg* updates
        if t >= repeat.thisN == 5 and exitMsg.status == NOT_STARTED:
            # keep track of start time/frame for later
            exitMsg.tStart = t  # underestimates by a little under one frame
            exitMsg.frameNStart = frameN  # exact frame index
            exitMsg.setAutoDraw(True)
        
        # *contKeyPress* updates
        if t >= 0.0 and contKeyPress.status == NOT_STARTED:
            # keep track of start time/frame for later
            contKeyPress.tStart = t  # underestimates by a little under one frame
            contKeyPress.frameNStart = frameN  # exact frame index
            contKeyPress.status = STARTED
            # keyboard checking is just starting
            contKeyPress.clock.reset()  # now t=0
            event.clearEvents(eventType='keyboard')
        if contKeyPress.status == STARTED:
            theseKeys = event.getKeys(keyList=['space'])
            
            # check for quit:
            if "escape" in theseKeys:
                endExpNow = True
            if len(theseKeys) > 0:  # at least one key was pressed
                contKeyPress.keys = theseKeys[-1]  # just the last key pressed
                contKeyPress.rt = contKeyPress.clock.getTime()
                # a response ends the routine
                continueRoutine = False
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in nextTrialComponents:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # check for quit (the Esc key)
        if endExpNow or event.getKeys(keyList=["escape"]):
            core.quit()
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    #-------Ending Routine "nextTrial"-------
    for thisComponent in nextTrialComponents:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    # check responses
    if contKeyPress.keys in ['', [], None]:  # No response was made
       contKeyPress.keys=None
    # store data for repeat (TrialHandler)
    repeat.addData('contKeyPress.keys',contKeyPress.keys)
    if contKeyPress.keys != None:  # we had a response
        repeat.addData('contKeyPress.rt', contKeyPress.rt)
    # the Routine "nextTrial" was not non-slip safe, so reset the non-slip timer
    routineTimer.reset()
    thisExp.nextEntry()
    
# completed 6 repeats of 'repeat'

# get names of stimulus parameters
if repeat.trialList in ([], [None], None):  params = []
else:  params = repeat.trialList[0].keys()
# save data for this loop
repeat.saveAsExcel(filename + '.xlsx', sheetName='repeat',
    stimOut=params,
    dataOut=['n','all_mean','all_std', 'all_raw'])



win.close()
core.quit()
