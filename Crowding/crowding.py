from psychopy import visual, core, event, data
from psychopy import gui
#from win32api import GetSystemMetrics
import pyglet
from pyglet.window import key
#import xlwt
import os
import openpyxl as pyxl
import random
import string
import random

s = False
 
#----------------------------------------------------------------------#
#STEP 000: DEFINE PARAMS
#----------------------------------------------------------------------#
unscaled_distance = [15,45,80]  #letter spacing in pixels
d_scaler = 0.05                 #distance scaler to be multiplied by fontsize and unscaled distance
type = ['a','xa','ax','xax']    #alone,foveal,peripheral,middle
textfontsize = 20               #for instructions
flankercolor = 'white'          #color for the flankers
waittrial = 0.2                 #berief delay before each trial after subject's keypress
waitstim =  1.0                 #duration for the second presentation of the stimulus
#----------------------------------------------------------------------#
#STEP 00:INITIALIZE EMPTY LISTS TO STORE SOME DATA
#----------------------------------------------------------------------#
result=[]
response=[]
targets=[]
sides=[]
#----------------------------------------------------------------------#
#STEP 0: DEFINE NECESSARY FUNCTIONS
#----------------------------------------------------------------------#
# this function writes output to a file

def write_file(filename):
    global result
    book = pyxl.Workbook()
    ws = book.active
    ws.title = "Sheet 1"
    # write experiment information
    ws.cell(row=1,column=1).value = "ExperimentName:%s"%name
    ws.cell(row=2,column=1).value = "SubjectID:%s"%ID
    ws.cell(row=3,column=1).value = "SessionNumber:%s"%session
    # write column headers
    headers = ['Trial','Eccentricity (px)','Fontsize','LetterSpacing','Condition','LetterPair','Target','Side','Response','Result']
    for col,header in enumerate(headers):
        ws.cell(row=4,column=col+1).value = header
    # write per-trial data
    for r in range(0,len(result)):
        ws.cell(r+5, 1).value =  r+1
        ws.cell(r+5, 2).value = int(ecc)
        ws.cell(r+5, 3).value = fontsize
        ws.cell(r+5, 4).value = c[r][1]
        ws.cell(r+5, 5).value = c[r][0]
        ws.cell(r+5, 6).value = ' '.join(stim[r])
        ws.cell(r+5, 7).value = targets[r]
        ws.cell(r+5, 8).value = sides[r]
        ws.cell(r+5, 9).value = response[r].upper()
        ws.cell(r+5, 10).value = result[r]

    # create directory and/or save file
    if not os.path.exists('data'):
        os.makedirs('data')
    book.save('data/'+filename+'.xlsx')

#this function draws fixation point
def drawfixation():
    global win
    fix=visual.Circle(win, fillColor='white',lineColor='white',radius=5, edges=32,pos=(0,0))
    #fix=visual.Circle(win, fillColor='black',lineColor='black',radius=5, edges=32,pos=(0,0))
    return fix

#this function generates random stimulus (letters) for a given trial
def generate_id(current_cond,lettercase):
    if lettercase == 'Upper': return random.sample(string.ascii_uppercase,len(current_cond))
    else:  return random.sample(string.ascii_lowercase,len(current_cond))

#this function generates a random order of the conditions, with a given repetition:type=condition type (a,ax,xa,xax), 
#distance=distance between target and flankers,nr_rep=number of repetitions for each condition X distance (4X3=12), e.g. if nr_rep=2, then total # of trials=2*12=24
def generate_cond(type,distance,nr_rep):
    cond = [[config,int(dist)] for config in type for dist in distance]*nr_rep
    rand_cond = random.sample(cond,len(cond))
    return rand_cond

#this function generates stimuli for all trials: c=randomized condition list
def generate_stim(c,lettercase):
    stim = [generate_id(k[0],lettercase) for k in c]
    return stim

#this function gets current target and distance for a given trial: c=randomized condition list, stim=letter pair
def get_currents(c,stim):
    current_dist = c[1] #get the current distance between target and flankers
    if c[0]=='a':
        current_target = stim[0]
    elif c[0]=='ax':
        current_target = stim[0]
    elif c[0]=='xa':
        current_target = stim[1]
    elif c[0]=='xax':
        current_target = stim[1]
    return current_target,current_dist

#this function decidea side of presentation
def get_side():
    global sides
    if random.random()<0.5:side=-1
    else: side = 1
    sides.append(side)
    return side

#this function draws stimulus for a given trial: c=randomized condition list, stim=letter pair,dist=distance between target and the flankers,
#ecc=eccentricity of the target,side=side of the presentation(left or right)
def draw_stim(c,stim,dist,ecc,side):
    global win
    global fontsize,targetcolor,flankercolor
    #draw stim
    if c[0] == 'a': # check the condition type and draw accordingly
        t=visual.TextStim(win,text=stim[0],color=targetcolor, height=fontsize, pos=(ecc*side,0)) #draw target
        t.draw()
    elif c[0] == 'ax':
        t=visual.TextStim(win,text=stim[0],color=targetcolor, height=fontsize, pos=(ecc*side,0)) #draw target
        f=visual.TextStim(win,text=stim[1],color=flankercolor, height=fontsize, pos=((ecc+dist)*side,0)) #draw flanker
        t.draw()
        f.draw()
    elif c[0] == 'xa':
        t=visual.TextStim(win,text=stim[1],color=targetcolor, height=fontsize, pos=(ecc*side,0)) #draw target
        f=visual.TextStim(win,text=stim[0],color=flankercolor, height=fontsize, pos=((ecc-dist)*side,0)) #draw flanker
        t.draw()
        f.draw()
    elif c[0] == 'xax':
        t=visual.TextStim(win,text=stim[1],color=targetcolor, height=fontsize, pos=(ecc*side,0)) #draw target
        f=visual.TextStim(win,text=stim[0],color=flankercolor, height=fontsize, pos=((ecc-dist)*side,0)) #draw flanker
        f2=visual.TextStim(win,text=stim[2],color=flankercolor, height=fontsize, pos=((ecc+dist)*side,0)) #draw flanker
        t.draw()
        f.draw()
        f2.draw()
    win.flip()

#this function collects response for a given trial
def collect_response():
    global win
    global result,targets,response,current_target
    # first show instruction
    show_instruction()
    win.flip()
    #then get response 
    thisResp=[]
    while thisResp==[]:
        allKeys=event.waitKeys()
        for thisKey in allKeys: #allow for cancellation
            if thisKey == 'escape':
                print('user cancelled')
                core.quit() #abort the experiment
            elif thisKey==current_target or thisKey==current_target.lower(): thisResp=1 # correct response
            else: thisResp=-1 #incorrect response
    result.append(thisResp) #add current result to the result list
    if thisKey == 'space': thisKey=9
    response.append(thisKey) #add current response to the response list
    targets.append(current_target) #add current target to the target list
    event.clearEvents()

# this function shows trial instruction: if arg is given, function draws first instruction, otherwise draws the second instruction.
def show_instruction(idx=None,trial_nr=None,trials=None):
    global win
    drawfixation().draw()
    if idx==None:
        inst=visual.TextStim(win,text='Press key of the target letter.',color='white', height=textfontsize, pos=(0,100)) #draw instruction text
    elif idx==1:
        inst=visual.TextStim(win,text='The previous display was:',color='white', height=textfontsize, pos=(0,100)) #draw instruction text
    elif idx==2:
        inst=visual.TextStim(win,text='Trial %d out of %d trials. Press any key to start.'%(trial_nr,trials),color='white', height=textfontsize, pos=(0,100)) #draw instruction text
    inst.draw()

#this function shows introduction text to the experiment
def show_introduction():
    global win
    intro=visual.TextStim(win,\
        text='On each trial a target letter will appear either alone or with one or two flanking letters.\
            \nThe flanking letters will always be white.\
            \nThe target letter color was already selected by you on the initial screen.\
            \n\
            \nKeep looking at the central dot.\
            \nDO NOT look toward the letters.\
            \nPress a key to start the trial.\
            \nAt the end of the trial press a key to identify the target letter.\
            \nThe correct answer will be shown after the trial.',
        color='white', height=textfontsize, anchorHoriz='center', anchorVert='center') #draw introduction text
    intro.draw()
    win.flip()
    while True:
        if len(event.getKeys())>0: break
    event.clearEvents()

#----------------------------------------------------------------------#
#STEP 1: INITIALIZE A DIALOG BOX
#----------------------------------------------------------------------#
myDlg = gui.Dlg(title="Crowding Experiment",size=(1, 1))
myDlg.addField('Experiment Name:','CROWDING')
myDlg.addField('Subject ID:','XX')
myDlg.addField('Session Number:', )
myDlg.addField('Eccentricity (pixels):', )
myDlg.addField('Letter Font Size:',26)
myDlg.addField('Letter Case:','Upper')
myDlg.addField('Number of Trials per Condition',4)
myDlg.addField('Duration (ms)',200)
myDlg.addField('Target Color','red')
myDlg.show()
#----------------------------------------------------------------------#
#STEP 2: LOAD IN INPUT FROM THE DIALOG BOX
#----------------------------------------------------------------------#
if myDlg.OK:  # then the user pressed OK
    thisInfo = myDlg.data
    name=thisInfo[0]
    ID=thisInfo[1]
    session=thisInfo[2]
    ecc=int(thisInfo[3])
    fontsize=int(thisInfo[4])
    lettercase=thisInfo[5]
    nr_rep=thisInfo[6]
    duration=int(thisInfo[7])/1000 #get duration in miliseconds
    targetcolor=thisInfo[8]
#----------------------------------------------------------------------#
#STEP 3: DEFINE FILE TO WRITE DATA ON
#----------------------------------------------------------------------#
    file=''
    file+=name
    file+='_'+ID
    file+='_'+str(session)
    file+='_'+data.getDateStr()
    print(file)
#----------------------------------------------------------------------#
#STEP 4: OPEN A WINDOW
#----------------------------------------------------------------------#
    win = visual.Window(monitor='testMonitor',color='black',allowGUI=True,units='pix',fullscr=True)
#----------------------------------------------------------------------#
#STEP 5: GET STIMULUS LIST FOR THE CURRENT BLOCK
#----------------------------------------------------------------------#
    distance = [d*fontsize*d_scaler for d in unscaled_distance] # scale the distance between target and flankers depending on the current fontsize
    c=generate_cond(type,distance,nr_rep) #create a random order condition list using the number of repetitions
    stim=generate_stim(c,lettercase) #generate stimulus list
    trials=len(c) #get the number of trials
#----------------------------------------------------------------------#
#STEP 6: SHOW INTRODUCTION AND WAIT FOR A KEY PRESS TO START THE TRIAL.
#        THEN FOR EACH TRIAL:
#           A.DETERMINE THE SIDE OF PRESENTATION: LEFT (-1) OR RIGHT (1)
#           B.DRAW FIXATION POINT
#           C.GET CURRENT TARGET AND DISTANCE BETWEEN TARGET AND THE FLANKERS 
#           D.DRAW AND PRESENT STIMULUS FOR GIVEN DURATION
#           E.COLLECT RESPONSE
#           F.THEN PRESENT STIMULUS ONE MORE TIME FOR A LONGER PERIOD (1 SEC)
#           G.CLEAR EVENTS BEFORE MOVING TO THE NEXT TRIAL
#----------------------------------------------------------------------#
    show_introduction()
    for i in range(0,trials):
        if s==True:
            break;
        else:
            show_instruction(2,i+1,trials) # show start trial instruction
            win.flip()
            allKeys=event.waitKeys() #check for key press to start the trial
            for thisKey in allKeys:
                if thisKey=='escape': # if user pressed escape, quit the experiment
                    print('user cancelled')
                    core.quit()
                else:
                    core.wait(waittrial) #a brief delay before starting the actual trial
                    side=get_side() #determine the side of presentation
                    drawfixation().draw() #draw fixation
                    current_target,current_dist=get_currents(c[i],stim[i]) #current target and the spacing between target and the flankers
                    draw_stim(c[i],stim[i],current_dist,ecc,side) #draw stimulus
                    core.wait(duration) #wait for stimulus presentation
                    collect_response() #show response screen and collect response
                    show_instruction(1) #draw stimulus one more time for a longer duration
                    drawfixation().draw() #draw fixation
                    draw_stim(c[i],stim[i],current_dist,ecc,side) #draw stimulus for longer time
                    core.wait(waitstim)
            event.clearEvents()
#----------------------------------------------------------------------#
#STEP 7: WRITE DATA TO A FILE
#----------------------------------------------------------------------#
    write_file(file) #write data to a file
else:
    print('user cancelled')

