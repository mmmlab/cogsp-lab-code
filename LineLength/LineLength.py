from psychopy import visual, core, event, data
from psychopy import gui
# from win32api import GetSystemMetrics
import pyglet
from pyglet.window import key
from numpy import clip
#import xlwt
import openpyxl as pyxl
import random
import os

MAX_COMP_LENGTH = 0.5       # in height-normalized units
MIN_COMP_LENGTH = 0.0002    # in height-normalized units
LARGE_INCREMENT = 10        # in pixels
SMALL_INCREMENT = 1         # in pixels

s = False
result=[]


## Class defining ML (Muller-Lyer) stimulus object
class MLStimulus:
    def __init__(self,stand_length,position,is_muller,color='White',width=2):
        # length of the 'standard' portion of the stimulus
        self.stand_length = stand_length
        # length of the 'comparison' portion of the stimulus
        self.comp_length = random.uniform(stand_length*0.8,stand_length*1.2)
        # horizontal position (of stimulus center)
        self.xpos = position[0]
        # vertical position (of stimulus center)
        self.ypos = position[1]
        # is the stimulus a Muller-Lyer stimulus (with chevrons)?
        self.is_muller = is_muller
        self.color = color
        self.width = width

    def increment_comparison(self,increment):
        new_length = self.comp_length+increment
        self.comp_length = clip(new_length,MIN_COMP_LENGTH,MAX_COMP_LENGTH)

    def make_line(self,startpos,endpos):
        line = visual.ShapeStim(win,vertices=[startpos,endpos],
                lineColor=self.color,lineWidth=self.width,fillColor=None)
        return line

    def draw_chevron(self,hpos,direction):
        height = 0.025 # chevron height (height-normalized units)
        if direction < 0: # chevron points left 
            top_part = self.make_line([hpos+height,self.ypos+height],[hpos,self.ypos])
            bottom_part = self.make_line([hpos,self.ypos],[hpos+height,self.ypos-height]) 
        elif direction > 0: # chevron points right
            top_part = self.make_line([hpos-height,self.ypos+height],[hpos,self.ypos])
            bottom_part = self.make_line([hpos,self.ypos],[hpos-height,self.ypos-height]) 
        elif direction == 0: # non-chevron (straight-line terminations)
            top_part = self.make_line([hpos,self.ypos+height],[hpos,self.ypos])
            bottom_part = self.make_line([hpos,self.ypos],[hpos,self.ypos-height]) 
        top_part.draw()
        bottom_part.draw()

    def draw(self):
        left_hpos = self.xpos-self.stand_length
        mid_hpos = self.xpos
        right_hpos = self.xpos+self.comp_length
        # create and draw main line
        main_line = self.make_line([left_hpos,self.ypos],[right_hpos,self.ypos])
        main_line.draw()
        # draw chevrons
        self.draw_chevron(left_hpos,-1*self.is_muller) # leftmost chevron
        self.draw_chevron(mid_hpos,1*self.is_muller) # middle chevron
        self.draw_chevron(right_hpos,-1*self.is_muller)


################################################################################
## Utility Function to Compute Screen Resolution and Pixel Scaling
def get_display_info():
    """
    gets effective fullscreen resolution from the os and creates a temporary
    window to compute the native hardware resolution and to determine whether
    there is any pixel scaling (i.e., as in MacOS 'Retina' displays)

    returns a tuple consisting of: horizontal resolution, vertical resolution, 
    and pixel scaling factor
    """
    # 1. Get full display resolution (as reported by OS)
    # note: code below assumes either single or twin monitor configuration
    screen = pyglet.canvas.Display().get_default_screen()
    os_width = screen.width
    os_height = screen.height
    # 2. Create fullscreen Psychopy window and get actual hardware resolution
    testwin = visual.Window(monitor='testMonitor',color='black',allowGUI=True,\
        units='pix',size=(os_width,os_height),fullscr=True)
    hard_width,hard_height = testwin.size
    # 3. Compute the ratio of hard_width to os_width to determine whether we're
    #    using a HiDPI display (i.e., one with pixel scaling).
    pixel_scaling = hard_width/os_width
    # 4. Exit the window
    testwin.close()
    return os_width,os_height,pixel_scaling
################################################################################

def write_file(filename,length,mueller):
    global result
    book = pyxl.Workbook()
    ws = book.active
    ws.title = "Sheet 1"
    # write experiment information
    ws.cell(row=1,column=1).value = "ExperimentName:%s"%name
    ws.cell(row=2,column=1).value = "SubjectID:%s"%ID
    ws.cell(row=3,column=1).value = "SessionNumber:%s"%session
    # write column headers
    headers = ['Trial','StandardLength','Mueller-Lyer?','ComparisonLength']
    for col,header in enumerate(headers):
        ws.cell(row=4,column=col+1).value = header
    # write per-trial data
    for i in range(0,len(result)):
        ws.cell(i+5, 1).value = i+1
        ws.cell(i+5, 2).value = length
        ws.cell(i+5, 3).value = mueller
        ws.cell(i+5, 4).value =result[i]
    # create data directory and save data file
    if not os.path.exists('data'):
        os.makedirs('data')
    book.save('data/'+filename+'.xlsx') 

def show_stimulus(length,is_muller):
    global win
    global textbox
    pos_jitter = random.uniform(-0.05, 0.05)
    winHeight = win.size[1]
    s_length=length/winHeight
    stimulus = MLStimulus(s_length,[pos_jitter,0],is_muller=is_muller)
    instruction_text = """This is trial %s
        The left line is the standard, and the right line is the comparison.
        Press J to INCREASE the comparison by 10 pixels
        Press F to DECREASE the comparison by 10 pixels.
        Press H to INCREASE the comparison by 1 pixel
        Press G to DECREASE the comparison by 1 pixel.
        Press SPACEBAR when done with this trial.
        Press Esc to end the session early"""%str(trial_count+1)
    textbox.setText(instruction_text)
    textbox.draw()
    stimulus.draw()
    win.flip()
    while True:
        keylist=event.getKeys()
        increment = 0
        if 'space'in keylist:
            global result
            comparison_length = stimulus.comp_length*winHeight
            result.append(int(comparison_length))
            print('insert new length to c_length list')
            break 
        elif keyState[key.J]:
            increment = LARGE_INCREMENT/winHeight
        elif keyState[key.H]:
            increment = SMALL_INCREMENT/winHeight
        elif keyState[key.F]:
            increment = -LARGE_INCREMENT/winHeight
        elif keyState[key.G]:
            increment = -SMALL_INCREMENT/winHeight
        elif 'escape'in keylist:
            print('set escape signal')
            global s
            s=True
            break
        else:
            pass
        stimulus.increment_comparison(increment)
        textbox.draw()
        stimulus.draw()
        win.flip()

def show_instruction():
    global win
    global textbox
    global trial_count
    showbox=visual.TextStim(win,text='PRESS ANY KEY TO START', font='ARIAL',
        color=[0,0,0],pos=(0.0,-0.25),units='height',colorSpace='rgb255')
    showbox.height = 0.7
    showbox.wrapWidth = 1
    showbox.draw()
    textbox.draw()
    win.flip()
    while True:
        if len(event.getKeys())>0: break
    event.clearEvents()
 
    
myDlg = gui.Dlg(title="Line Length",size=(1, 1))
myDlg.addField('ExperimentName:','LINE')
myDlg.addField('SubjectID:','xx')
myDlg.addField('SessionNumber:',1)
myDlg.addField('StandardLineLength(in pixels):',200)
myDlg.addField('NumberofTrials',30)
myDlg.addField('Mueller-Lyer(N;Y)')
myDlg.show()
if myDlg.OK:  # then the user pressed OK
    thisInfo = myDlg.data
    name=thisInfo[0]
    ID=thisInfo[1]
    session=thisInfo[2]
    length=thisInfo[3]
    trials=thisInfo[4]
    mueller=thisInfo[5]
    
    
    
    file=''
    file+=name
    file+='_'
    file+=ID
    file+='_'
    file+=str(session)
    file+='_'+data.getDateStr()

    print(file)
    keyState=key.KeyStateHandler()
    wwidth,wheight,px_scale = get_display_info()
    win = visual.Window([wwidth,wheight],monitor='Monitor',allowGUI=True,units='height',fullscr=True)
    instruction_text = """The left line is the standard, and the right line is the comparison.
        Press J to INCREASE the comparison by 10 pixels
        Press F to DECREASE the comparison by 10 pixels.\nPress H to INCREASE the comparison by 1 pixel
        Press G to DECREASE the comparison by 1 pixel.
        Press SPACEBAR when done with this trial.
        Press Esc to end the session early"""
    textbox=visual.TextStim(win,text=instruction_text, font='Arial',
        color=[0,0,0],pos=(0.0,0.25),units='height',colorSpace='rgb255')
    textbox.height = (0.025)
    textbox.wrapWidth = 1
    win.winHandle.push_handlers(keyState)
    show_instruction()
    trial_count = 0
    while trial_count < trials :
        if s==True:
            break;
        if mueller in 'Yy':
            show_stimulus(length,True)
        else:
            show_stimulus(length,False)
        trial_count += 1
    write_file(file,length,mueller)
else:
    print('user cancelled')